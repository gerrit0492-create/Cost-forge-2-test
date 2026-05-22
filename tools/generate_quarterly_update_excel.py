"""
Quarterly Cost Update Excel Generator
======================================
Generates a fully pre-loaded Excel workbook for the quarterly review cycle.
Every 3 months, run this script to produce a fresh workbook with current data.
Fill in the "New Price / Rate" columns, then run import_from_quarterly_excel.py
to write the changes back to the CSV databases.

Usage:
    python tools/generate_quarterly_update_excel.py
    # -> creates:  data/quarterly_update_<YYYY-Qn>.xlsx

Requirements:
    pip install openpyxl
"""

import os
import sys
import math
import csv
import json
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError:
    sys.exit("openpyxl not found. Run: pip install openpyxl")

# ─── Paths ────────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR  = REPO_ROOT / "data"

def data(fn): return DATA_DIR / fn

# ─── Colour palette ───────────────────────────────────────────────────────────
C_NAVY      = "0D3B66"
C_BLUE      = "1565C0"
C_TEAL      = "00838F"
C_GREEN     = "2E7D32"
C_AMBER     = "F57F17"
C_RED       = "C62828"
C_LIGHT_BG  = "F5F7FA"
C_HEADER_BG = "1565C0"
C_STRIPE    = "EBF3FF"
C_INPUT_BG  = "FFFDE7"   # pale yellow  – editable cells
C_LOCKED_BG = "ECEFF1"   # pale grey    – formula / locked cells
C_WHITE     = "FFFFFF"
C_INCREASE  = "FFCDD2"   # light red    – price went up
C_DECREASE  = "C8E6C9"   # light green  – price went down
C_NEUTRAL   = "F5F5F5"

# ─── Shared style helpers ─────────────────────────────────────────────────────
def hfont(bold=True, size=11, color=C_WHITE):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def bfont(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

thin = Side(border_style="thin", color="BDBDBD")
thick = Side(border_style="medium", color=C_NAVY)

def border_all():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def border_header():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def euro(ws, cell_ref):
    ws[cell_ref].number_format = '€#,##0.00'

def pct(ws, cell_ref):
    ws[cell_ref].number_format = '+0.00%;-0.00%;0.00%'

# ─── Quarter helpers ──────────────────────────────────────────────────────────
def current_quarter(d: date):
    q = (d.month - 1) // 3 + 1
    return d.year, q

def quarter_label(year, q):
    return f"{year}-Q{q}"

def next_quarter_start(d: date):
    year, q = current_quarter(d)
    if q == 4:
        return date(year + 1, 1, 1)
    return date(year, q * 3 + 1, 1)

# ─── CSV loaders ──────────────────────────────────────────────────────────────
def load_csv(path, comment_char="#"):
    rows = []
    if not path.exists():
        return [], []
    with open(path, newline="", encoding="utf-8") as f:
        lines = [l for l in f if not l.strip().startswith(comment_char)]
    reader = csv.DictReader(lines)
    for row in reader:
        rows.append(row)
    headers = list(reader.fieldnames) if reader.fieldnames else []
    return headers, rows

# ─── Sheet: Instructions ──────────────────────────────────────────────────────
def build_instructions(wb, today, year, q):
    ws = wb.create_sheet("📋 Instructions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70
    ws.row_dimensions[1].height = 6

    def row(r, label, value, bold_val=False):
        ws.row_dimensions[r].height = 22
        ws[f"B{r}"].value = label
        ws[f"B{r}"].font = Font(bold=True, size=10, name="Calibri", color=C_NAVY)
        ws[f"B{r}"].alignment = left()
        ws[f"C{r}"].value = value
        ws[f"C{r}"].font = Font(bold=bold_val, size=10, name="Calibri")
        ws[f"C{r}"].alignment = left()

    ws.merge_cells("B2:C2")
    ws["B2"].value = "⚙️  COST FORGE — Quarterly Cost Update Workbook"
    ws["B2"].font = Font(bold=True, size=16, name="Calibri", color=C_WHITE)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 38

    ws.merge_cells("B3:C3")
    ws["B3"].value = f"Generated for:  {quarter_label(year, q)}   |   Date: {today.strftime('%d %B %Y')}"
    ws["B3"].font = Font(italic=True, size=11, name="Calibri", color=C_WHITE)
    ws["B3"].fill = fill(C_BLUE)
    ws["B3"].alignment = center()
    ws.row_dimensions[3].height = 24

    ws.row_dimensions[5].height = 26
    ws.merge_cells("B5:C5")
    ws["B5"].value = "🗂  WORKBOOK TABS OVERVIEW"
    ws["B5"].font = Font(bold=True, size=12, name="Calibri", color=C_WHITE)
    ws["B5"].fill = fill(C_TEAL)
    ws["B5"].alignment = center()

    tabs = [
        ("📋 Instructions",       "This page — overview, colour key, step-by-step guide."),
        ("📅 Q-Schedule",          "Quarterly review calendar showing past and upcoming cycles."),
        ("💎 Materials DB",        "Update material prices (€/kg). Yellow cells = editable."),
        ("⚙️ Process Rates",       "Update machine / labour rates and overhead %. Yellow = editable."),
        ("🏢 Supplier Quotes",     "Update supplier quote prices, lead times and expiry dates."),
        ("📊 Market Adjustments",  "Apply % or factor adjustments across commodity families."),
        ("💰 Cost Impact",         "Live cost impact of your changes on existing BOM lines."),
        ("📤 Export Preview",      "Copy-ready CSV rows for all changed records."),
    ]
    for i, (tab, desc) in enumerate(tabs, start=6):
        ws.row_dimensions[i].height = 20
        ws[f"B{i}"].value = tab
        ws[f"B{i}"].font = Font(bold=True, size=10, name="Calibri", color=C_NAVY)
        ws[f"B{i}"].alignment = left()
        ws[f"C{i}"].value = desc
        ws[f"C{i}"].font = bfont(size=10)
        ws[f"C{i}"].alignment = left()
        if i % 2 == 0:
            ws[f"B{i}"].fill = fill(C_STRIPE)
            ws[f"C{i}"].fill = fill(C_STRIPE)

    ws.row_dimensions[15].height = 26
    ws.merge_cells("B15:C15")
    ws["B15"].value = "📌  STEP-BY-STEP PROCESS"
    ws["B15"].font = Font(bold=True, size=12, name="Calibri", color=C_WHITE)
    ws["B15"].fill = fill(C_TEAL)
    ws["B15"].alignment = center()

    steps = [
        ("Step 1",  "Open the 💎 Materials DB tab. Enter new prices in the yellow 'New Price (€/kg)' column."),
        ("Step 2",  "Open ⚙️ Process Rates. Update machine rates, labour rates, and overheads as needed."),
        ("Step 3",  "Open 🏢 Supplier Quotes. Enter new quoted prices and update expiry dates to +90 days."),
        ("Step 4",  "Open 📊 Market Adjustments. Enter any commodity % adjustments for this quarter."),
        ("Step 5",  "Review the 💰 Cost Impact tab to see the total € and % change on open BOM lines."),
        ("Step 6",  "Run the import script:  python tools/import_from_quarterly_excel.py"),
        ("Step 7",  "The script writes all changes back to the CSV databases and logs a history entry."),
        ("Step 8",  "Commit & push the updated data/  folder. Done — next review in 90 days!"),
    ]
    for i, (step, desc) in enumerate(steps, start=16):
        ws.row_dimensions[i].height = 22
        ws[f"B{i}"].value = step
        ws[f"B{i}"].font = Font(bold=True, size=10, name="Calibri", color=C_WHITE)
        ws[f"B{i}"].fill = fill(C_BLUE if i % 2 == 0 else C_NAVY)
        ws[f"B{i}"].alignment = center()
        ws[f"C{i}"].value = desc
        ws[f"C{i}"].font = bfont(size=10)
        ws[f"C{i}"].alignment = left()

    ws.row_dimensions[25].height = 26
    ws.merge_cells("B25:C25")
    ws["B25"].value = "🎨  COLOUR KEY"
    ws["B25"].font = Font(bold=True, size=12, name="Calibri", color=C_WHITE)
    ws["B25"].fill = fill(C_TEAL)
    ws["B25"].alignment = center()

    colours = [
        (C_INPUT_BG, "Yellow — Data entry cell (you fill this in)"),
        (C_LOCKED_BG,"Grey — Auto-calculated (do not edit)"),
        (C_INCREASE,  "Light red — Price has increased vs. current value"),
        (C_DECREASE,  "Light green — Price has decreased vs. current value"),
        ("FFEE58",    "Amber — Quote expiring within 30 days"),
        (C_RED,       "Red — Quote already expired"),
    ]
    for i, (hex_c, desc) in enumerate(colours, start=26):
        ws.row_dimensions[i].height = 20
        ws[f"B{i}"].fill = fill(hex_c)
        ws[f"B{i}"].value = "  " + desc.split("—")[0].strip()
        ws[f"B{i}"].font = Font(size=10, name="Calibri")
        ws[f"B{i}"].alignment = left()
        ws[f"C{i}"].value = desc.split("—")[1].strip()
        ws[f"C{i}"].font = bfont(size=10)
        ws[f"C{i}"].alignment = left()

# ─── Sheet: Q-Schedule ────────────────────────────────────────────────────────
def build_schedule(wb, today, year, q):
    ws = wb.create_sheet("📅 Q-Schedule")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFG", [18, 18, 18, 18, 22, 30]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:G2")
    ws["B2"].value = "📅  Quarterly Review Schedule"
    ws["B2"].font = Font(bold=True, size=14, name="Calibri", color=C_WHITE)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 32

    headers = ["Quarter", "Start Date", "Due Date", "Status", "Completed By", "Notes"]
    for ci, h in enumerate(headers, start=2):
        col = get_column_letter(ci)
        ws[f"{col}4"].value = h
        ws[f"{col}4"].font = hfont(size=10)
        ws[f"{col}4"].fill = fill(C_HEADER_BG)
        ws[f"{col}4"].alignment = center()
        ws[f"{col}4"].border = border_all()
    ws.row_dimensions[4].height = 20

    ref_year, ref_q = current_quarter(today)
    quarters = []
    y, qn = ref_year, ref_q
    for _ in range(4):
        qn -= 1
        if qn == 0:
            qn = 4
            y -= 1
    for i in range(8):
        start = date(y, (qn - 1) * 3 + 1, 1)
        due   = start + timedelta(days=89)
        quarters.append((y, qn, start, due))
        qn += 1
        if qn > 4:
            qn = 1
            y += 1

    for ri, (qy, qq, start, due) in enumerate(quarters, start=5):
        r = ri
        ws.row_dimensions[r].height = 20
        label = quarter_label(qy, qq)
        is_current = (qy == ref_year and qq == ref_q)
        is_past    = due < today

        if is_current:
            status = "🔄 Current"
            row_fill = fill("E3F2FD")
        elif is_past:
            status = "✅ Completed"
            row_fill = fill(C_STRIPE)
        else:
            status = "⏳ Upcoming"
            row_fill = fill(C_NEUTRAL)

        data = [label, start.strftime("%d %b %Y"), due.strftime("%d %b %Y"), status, "", ""]
        for ci, val in enumerate(data, start=2):
            col = get_column_letter(ci)
            ws[f"{col}{r}"].value = val
            ws[f"{col}{r}"].font = Font(bold=is_current, size=10, name="Calibri",
                                        color=C_NAVY if is_current else "000000")
            ws[f"{col}{r}"].fill = row_fill
            ws[f"{col}{r}"].alignment = center()
            ws[f"{col}{r}"].border = border_all()
            if ci >= 6:
                ws[f"{col}{r}"].fill = fill(C_INPUT_BG) if not is_past else row_fill

    next_start = next_quarter_start(today)
    ws.row_dimensions[14].height = 8
    ws.merge_cells("B15:G15")
    ws["B15"].value = f"▶  Next update due:  {next_start.strftime('%d %B %Y')}  ({(next_start - today).days} days from today)"
    ws["B15"].font = Font(bold=True, size=12, name="Calibri", color=C_WHITE)
    ws["B15"].fill = fill(C_AMBER)
    ws["B15"].alignment = center()
    ws.row_dimensions[15].height = 28

# ─── Sheet: Materials DB ──────────────────────────────────────────────────────
def build_materials(wb):
    ws = wb.create_sheet("💎 Materials DB")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    col_widths = {"A":4,"B":14,"C":36,"D":16,"E":14,"F":16,"G":14,"H":12,"I":24}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("B2:I2")
    ws["B2"].value = "💎  Materials Database — Quarterly Price Update"
    ws["B2"].font = hfont(size=13)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:I3")
    ws["B3"].value = ("Enter new prices in the yellow column (F). Leave blank to keep current price. "
                      "The % Change column updates automatically.")
    ws["B3"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws["B3"].fill = fill(C_LIGHT_BG)
    ws["B3"].alignment = left()
    ws.row_dimensions[3].height = 20

    headers = ["Material ID","Description","Commodity","Current Price (€/kg)",
               "New Price (€/kg)","Change %","Change (€)","Status / Notes"]
    cols = list("BCDEFGHI")
    for ci, (h, col) in enumerate(zip(headers, cols)):
        ws[f"{col}4"].value = h
        ws[f"{col}4"].font = hfont(size=10)
        ws[f"{col}4"].fill = fill(C_HEADER_BG)
        ws[f"{col}4"].alignment = center()
        ws[f"{col}4"].border = border_all()
    ws.row_dimensions[4].height = 22

    _, mats = load_csv(data("materials_db.csv"))
    dv_status = DataValidation(type="list",
                               formula1='"✅ Updated,⏸ No Change,⚠️ Review Needed,❌ Remove"',
                               allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_status)

    for i, mat in enumerate(mats, start=5):
        r = i
        ws.row_dimensions[r].height = 20
        stripe = fill(C_STRIPE) if i % 2 == 0 else fill(C_WHITE)

        cur_price = float(mat.get("price_eur_per_kg", 0) or 0)

        row_data = [
            ("B", mat.get("material_id",""),   False),
            ("C", mat.get("description",""),   False),
            ("D", mat.get("commodity",""),      False),
            ("E", cur_price,                   False),
        ]
        for col, val, editable in row_data:
            ws[f"{col}{r}"].value = val
            ws[f"{col}{r}"].font = bfont(size=10)
            ws[f"{col}{r}"].fill = stripe
            ws[f"{col}{r}"].alignment = left() if col in ("C","D") else right()
            ws[f"{col}{r}"].border = border_all()
            if col == "E":
                ws[f"{col}{r}"].number_format = "€#,##0.00"

        ws[f"F{r}"].fill = fill(C_INPUT_BG)
        ws[f"F{r}"].border = border_all()
        ws[f"F{r}"].number_format = "€#,##0.00"
        ws[f"F{r}"].alignment = right()

        ws[f"G{r}"].value = f'=IF(F{r}="","",IF(E{r}=0,"N/A",(F{r}-E{r})/E{r}))'
        ws[f"G{r}"].number_format = '+0.00%;-0.00%;0.00%'
        ws[f"G{r}"].fill = fill(C_LOCKED_BG)
        ws[f"G{r}"].border = border_all()
        ws[f"G{r}"].alignment = right()

        ws[f"H{r}"].value = f'=IF(F{r}="","",(F{r}-E{r}))'
        ws[f"H{r}"].number_format = '+€#,##0.00;-€#,##0.00;€0.00'
        ws[f"H{r}"].fill = fill(C_LOCKED_BG)
        ws[f"H{r}"].border = border_all()
        ws[f"H{r}"].alignment = right()

        ws[f"I{r}"].fill = fill(C_INPUT_BG)
        ws[f"I{r}"].border = border_all()
        ws[f"I{r}"].alignment = center()

    n = len(mats)
    if n > 0:
        dv_status.sqref = f"I5:I{5+n-1}"

    first_data = 5
    last_data  = 5 + n - 1
    ws.conditional_formatting.add(
        f"F{first_data}:F{last_data}",
        FormulaRule(formula=[f"AND(F{first_data}<>\"\",F{first_data}>E{first_data})"],
                    fill=fill(C_INCREASE))
    )
    ws.conditional_formatting.add(
        f"F{first_data}:F{last_data}",
        FormulaRule(formula=[f"AND(F{first_data}<>\"\",F{first_data}<E{first_data})"],
                    fill=fill(C_DECREASE))
    )

    sr = last_data + 2
    ws.merge_cells(f"B{sr}:E{sr}")
    ws[f"B{sr}"].value = f"📊  {n} materials loaded. Fill yellow cells to update prices."
    ws[f"B{sr}"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws[f"B{sr}"].fill = fill(C_LIGHT_BG)

# ─── Sheet: Process Rates ─────────────────────────────────────────────────────
def build_processes(wb):
    ws = wb.create_sheet("⚙️ Process Rates")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    col_widths = {
        "A":4,"B":16,"C":30,
        "D":14,"E":16,"F":14,"G":16,
        "H":13,"I":15,"J":13,"K":15,
        "L":12
    }
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("B2:L2")
    ws["B2"].value = "⚙️  Process Rates Database — Quarterly Update"
    ws["B2"].font = hfont(size=13)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:L3")
    ws["B3"].value = "Yellow cells are editable. Leave blank to keep the current value. % change columns auto-calculate."
    ws["B3"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws["B3"].fill = fill(C_LIGHT_BG)
    ws["B3"].alignment = left()
    ws.row_dimensions[3].height = 20

    headers = [
        "Process ID","Description",
        "Curr. Machine (€/h)","New Machine (€/h)",
        "Curr. Labour (€/h)","New Labour (€/h)",
        "Curr. Overhead %","New Overhead %",
        "Curr. Margin %","New Margin %",
        "Status"
    ]
    cols = list("BCDEFGHIJKL")
    for h, col in zip(headers, cols):
        ws[f"{col}4"].value = h
        ws[f"{col}4"].font = hfont(size=9)
        ws[f"{col}4"].fill = fill(C_HEADER_BG)
        ws[f"{col}4"].alignment = center()
        ws[f"{col}4"].border = border_all()
    ws.row_dimensions[4].height = 36

    _, procs = load_csv(data("processes_db.csv"))
    dv_status = DataValidation(type="list",
                               formula1='"✅ Updated,⏸ No Change,⚠️ Review Needed"',
                               allow_blank=True)
    ws.add_data_validation(dv_status)

    for i, proc in enumerate(procs, start=5):
        r = i
        ws.row_dimensions[r].height = 20
        stripe = fill(C_STRIPE) if i % 2 == 0 else fill(C_WHITE)

        machine  = float(proc.get("machine_rate_eur_h", 0) or 0)
        labour   = float(proc.get("labor_rate_eur_h",  0) or 0)
        overhead = float(proc.get("overhead_pct",      0) or 0)
        margin   = float(proc.get("margin_pct",        0) or 0)

        fixed_cols = [
            ("B", proc.get("process_id",""),  left()),
            ("C", proc.get("description",""), left()),
            ("D", machine,  right()),
            ("F", labour,   right()),
            ("H", overhead, right()),
            ("J", margin,   right()),
        ]
        for col, val, align in fixed_cols:
            ws[f"{col}{r}"].value = val
            ws[f"{col}{r}"].font = bfont(size=10)
            ws[f"{col}{r}"].fill = stripe
            ws[f"{col}{r}"].alignment = align
            ws[f"{col}{r}"].border = border_all()
            if col in ("D","F"):
                ws[f"{col}{r}"].number_format = "€#,##0.00"
            elif col in ("H","J"):
                ws[f"{col}{r}"].number_format = "0.0%"

        for edit_col in ("E","G","I","K"):
            ws[f"{edit_col}{r}"].fill = fill(C_INPUT_BG)
            ws[f"{edit_col}{r}"].border = border_all()
            ws[f"{edit_col}{r}"].alignment = right()
            if edit_col in ("E","G"):
                ws[f"{edit_col}{r}"].number_format = "€#,##0.00"
            else:
                ws[f"{edit_col}{r}"].number_format = "0.0%"

        ws[f"L{r}"].fill = fill(C_INPUT_BG)
        ws[f"L{r}"].border = border_all()
        ws[f"L{r}"].alignment = center()

    n = len(procs)
    if n > 0:
        dv_status.sqref = f"L5:L{5+n-1}"
    sr = 5 + n + 1
    ws.merge_cells(f"B{sr}:L{sr}")
    ws[f"B{sr}"].value = f"📊  {n} processes loaded."
    ws[f"B{sr}"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws[f"B{sr}"].fill = fill(C_LIGHT_BG)

# ─── Sheet: Supplier Quotes ───────────────────────────────────────────────────
def build_supplier_quotes(wb, today):
    ws = wb.create_sheet("🏢 Supplier Quotes")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    col_widths = {"A":4,"B":22,"C":14,"D":16,"E":16,"F":13,"G":16,
                  "H":14,"I":16,"J":12,"K":14,"L":12}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("B2:L2")
    ws["B2"].value = "🏢  Supplier Quotes — Quarterly Renewal"
    ws["B2"].font = hfont(size=13)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:L3")
    ws["B3"].value = ("Yellow cells are editable. "
                      "Rows highlighted red = expired; amber = expiring within 30 days.")
    ws["B3"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws["B3"].fill = fill(C_LIGHT_BG)
    ws["B3"].alignment = left()
    ws.row_dimensions[3].height = 20

    headers = ["Supplier","Material ID","Curr. Price (€/kg)","New Price (€/kg)",
               "Change %","Curr. Lead (days)","New Lead (days)",
               "Curr. Valid Until","New Valid Until","Preferred","Status"]
    cols = list("BCDEFGHIJKL")
    for h, col in zip(headers, cols):
        ws[f"{col}4"].value = h
        ws[f"{col}4"].font = hfont(size=9)
        ws[f"{col}4"].fill = fill(C_HEADER_BG)
        ws[f"{col}4"].alignment = center()
        ws[f"{col}4"].border = border_all()
    ws.row_dimensions[4].height = 36

    _, sq_rows = load_csv(data("supplier_quotes.csv"))
    _, q_rows  = load_csv(data("quotes.csv"))
    all_rows   = sq_rows + [r for r in q_rows
                             if not any(s["material_id"]==r["material_id"] and
                                        s["supplier"]==r["supplier"] for s in sq_rows)]

    default_new_expiry = (today + timedelta(days=91)).strftime("%Y-%m-%d")
    dv_status = DataValidation(type="list",
                               formula1='"✅ Renewed,⏸ No Change,❌ Expired / Remove"',
                               allow_blank=True)
    dv_pref = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_pref)

    for i, row in enumerate(all_rows, start=5):
        r = i
        ws.row_dimensions[r].height = 20
        stripe = fill(C_STRIPE) if i % 2 == 0 else fill(C_WHITE)

        cur_price = float(row.get("price_eur_per_kg", 0) or 0)
        cur_lead  = row.get("lead_time_days", "")
        cur_valid = row.get("valid_until", "")
        try:
            valid_date = date.fromisoformat(cur_valid)
            days_left  = (valid_date - today).days
        except Exception:
            valid_date = None
            days_left  = 999

        fixed = [
            ("B", row.get("supplier",""),   left()),
            ("C", row.get("material_id",""), center()),
            ("D", cur_price,                right()),
            ("G", cur_lead,                 center()),
            ("I", cur_valid,                center()),
        ]
        for col, val, align in fixed:
            ws[f"{col}{r}"].value = val
            ws[f"{col}{r}"].font = bfont(size=10)
            ws[f"{col}{r}"].fill = stripe
            ws[f"{col}{r}"].alignment = align
            ws[f"{col}{r}"].border = border_all()
            if col == "D":
                ws[f"{col}{r}"].number_format = "€#,##0.00"

        ws[f"E{r}"].fill = fill(C_INPUT_BG)
        ws[f"E{r}"].border = border_all()
        ws[f"E{r}"].number_format = "€#,##0.00"
        ws[f"E{r}"].alignment = right()

        ws[f"F{r}"].value = f'=IF(E{r}="","",IF(D{r}=0,"N/A",(E{r}-D{r})/D{r}))'
        ws[f"F{r}"].number_format = '+0.00%;-0.00%;0.00%'
        ws[f"F{r}"].fill = fill(C_LOCKED_BG)
        ws[f"F{r}"].border = border_all()
        ws[f"F{r}"].alignment = right()

        ws[f"H{r}"].fill = fill(C_INPUT_BG)
        ws[f"H{r}"].border = border_all()
        ws[f"H{r}"].alignment = center()

        ws[f"J{r}"].value = default_new_expiry
        ws[f"J{r}"].fill = fill(C_INPUT_BG)
        ws[f"J{r}"].border = border_all()
        ws[f"J{r}"].alignment = center()

        ws[f"K{r}"].value = int(row.get("preferred","1") or 1)
        ws[f"K{r}"].fill = fill(C_INPUT_BG)
        ws[f"K{r}"].border = border_all()
        ws[f"K{r}"].alignment = center()

        ws[f"L{r}"].fill = fill(C_INPUT_BG)
        ws[f"L{r}"].border = border_all()
        ws[f"L{r}"].alignment = center()

        if days_left < 0:
            for col in "BCDGI":
                ws[f"{col}{r}"].fill = fill("FFCDD2")
        elif days_left <= 30:
            for col in "BCDGI":
                ws[f"{col}{r}"].fill = fill("FFF9C4")

    n = len(all_rows)
    if n > 0:
        dv_status.sqref = f"L5:L{5+n-1}"
        dv_pref.sqref   = f"K5:K{5+n-1}"
    sr = 5 + n + 1
    ws.merge_cells(f"B{sr}:L{sr}")
    ws[f"B{sr}"].value = f"📊  {n} supplier quotes loaded."
    ws[f"B{sr}"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws[f"B{sr}"].fill = fill(C_LIGHT_BG)

# ─── Sheet: Market Adjustments ───────────────────────────────────────────────
def build_market(wb):
    ws = wb.create_sheet("📊 Market Adjustments")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    for col, w in zip("ABCDEFGH", [4,18,18,16,16,16,16,32]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:H2")
    ws["B2"].value = "📊  Market Adjustments — Commodity % / Factor Changes"
    ws["B2"].font = hfont(size=13)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:H3")
    ws["B3"].value = ("Either enter pct_change (e.g. 3 for +3%) OR factor (e.g. 1.03). "
                      "Leave both blank to skip. The script picks whichever is filled in.")
    ws["B3"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws["B3"].fill = fill(C_LIGHT_BG)
    ws["B3"].alignment = left()
    ws.row_dimensions[3].height = 22

    headers = ["Material ID","Commodity","Curr % Change","New % Change",
               "Curr Factor","New Factor","Note"]
    for ci, h in enumerate(headers, start=2):
        col = get_column_letter(ci)
        ws[f"{col}4"].value = h
        ws[f"{col}4"].font = hfont(size=10)
        ws[f"{col}4"].fill = fill(C_HEADER_BG)
        ws[f"{col}4"].alignment = center()
        ws[f"{col}4"].border = border_all()
    ws.row_dimensions[4].height = 22

    _, mf_rows = load_csv(data("market-factors.csv"))
    for i, row in enumerate(mf_rows, start=5):
        r = i
        ws.row_dimensions[r].height = 20
        stripe = fill(C_STRIPE) if i % 2 == 0 else fill(C_WHITE)

        fixed = [
            ("B", row.get("material_id",""),  left()),
            ("C", row.get("commodity",""),     left()),
            ("D", row.get("pct_change",""),    right()),
            ("F", row.get("factor",""),        right()),
            ("H", row.get("note",""),          left()),
        ]
        for col, val, align in fixed:
            try:
                val = float(val) if val else ""
            except Exception:
                pass
            ws[f"{col}{r}"].value = val
            ws[f"{col}{r}"].font = bfont(size=10)
            ws[f"{col}{r}"].fill = stripe
            ws[f"{col}{r}"].alignment = align
            ws[f"{col}{r}"].border = border_all()

        for edit_col in ("E", "G"):
            ws[f"{edit_col}{r}"].fill = fill(C_INPUT_BG)
            ws[f"{edit_col}{r}"].border = border_all()
            ws[f"{edit_col}{r}"].alignment = right()
            ws[f"{edit_col}{r}"].number_format = "0.00"

    n = len(mf_rows)
    sr = 5 + n + 1
    ws.merge_cells(f"B{sr}:H{sr}")
    ws[f"B{sr}"].value = f"📊  {n} market factor rows loaded."
    ws[f"B{sr}"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws[f"B{sr}"].fill = fill(C_LIGHT_BG)

# ─── Sheet: Cost Impact ───────────────────────────────────────────────────────
def build_cost_impact(wb):
    ws = wb.create_sheet("💰 Cost Impact")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    col_widths = {"A":4,"B":8,"C":14,"D":30,"E":12,"F":14,
                  "G":16,"H":16,"I":14,"J":12}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("B2:J2")
    ws["B2"].value = "💰  Cost Impact Analysis — Effect of Price Changes on Existing BOM Lines"
    ws["B2"].font = hfont(size=13)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:J3")
    ws["B3"].value = ("This sheet reads the current costs.csv. "
                      "'New Material Cost' uses the new price from 💎 Materials DB if filled in, "
                      "otherwise the current price. Update the Materials sheet first for live totals.")
    ws["B3"].font = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
    ws["B3"].fill = fill(C_LIGHT_BG)
    ws["B3"].alignment = left()
    ws.row_dimensions[3].height = 22

    headers = ["Line","Material ID","Description","Qty","Mass (kg)",
               "Curr Mat. Cost (€)","New Mat. Cost (€)","Delta (€)","Delta %"]
    cols = list("BCDEFGHIJ")
    for h, col in zip(headers, cols):
        ws[f"{col}4"].value = h
        ws[f"{col}4"].font = hfont(size=10)
        ws[f"{col}4"].fill = fill(C_HEADER_BG)
        ws[f"{col}4"].alignment = center()
        ws[f"{col}4"].border = border_all()
    ws.row_dimensions[4].height = 22

    _, costs = load_csv(data("costs.csv"))
    _, mats  = load_csv(data("materials_db.csv"))
    mat_row_map = {m["material_id"]: idx + 5 for idx, m in enumerate(mats)}

    for i, line in enumerate(costs, start=5):
        r = i
        ws.row_dimensions[r].height = 20
        stripe = fill(C_STRIPE) if i % 2 == 0 else fill(C_WHITE)

        mat_id   = line.get("material_id","")
        qty      = float(line.get("qty",1) or 1)
        mass_kg  = float(line.get("mass_kg",0) or 0)
        cur_mat  = float(line.get("material_cost",0) or 0)

        mat_row  = mat_row_map.get(mat_id)
        if mat_row:
            new_mat_formula = (
                f"=IF('💎 Materials DB'!F{mat_row}=\"\","
                f"E{r}*'💎 Materials DB'!E{mat_row},"
                f"E{r}*'💎 Materials DB'!F{mat_row})"
            )
        else:
            new_mat_formula = f"=G{r}"

        fixed_vals = [
            ("B", line.get("line_id",""),   center()),
            ("C", mat_id,                   center()),
            ("D", line.get("description",""), left()),
            ("E", qty,                      right()),
            ("F", mass_kg,                  right()),
        ]
        for col, val, align in fixed_vals:
            ws[f"{col}{r}"].value = val
            ws[f"{col}{r}"].font = bfont(size=10)
            ws[f"{col}{r}"].fill = stripe
            ws[f"{col}{r}"].alignment = align
            ws[f"{col}{r}"].border = border_all()

        ws[f"G{r}"].value = cur_mat
        ws[f"G{r}"].number_format = "€#,##0.00"
        ws[f"G{r}"].fill = stripe
        ws[f"G{r}"].font = bfont(size=10)
        ws[f"G{r}"].alignment = right()
        ws[f"G{r}"].border = border_all()

        ws[f"H{r}"].value = new_mat_formula
        ws[f"H{r}"].number_format = "€#,##0.00"
        ws[f"H{r}"].fill = fill(C_LOCKED_BG)
        ws[f"H{r}"].font = bfont(size=10)
        ws[f"H{r}"].alignment = right()
        ws[f"H{r}"].border = border_all()

        ws[f"I{r}"].value = f"=H{r}-G{r}"
        ws[f"I{r}"].number_format = '+€#,##0.00;-€#,##0.00;€0.00'
        ws[f"I{r}"].fill = fill(C_LOCKED_BG)
        ws[f"I{r}"].alignment = right()
        ws[f"I{r}"].border = border_all()

        ws[f"J{r}"].value = f'=IF(G{r}=0,"N/A",(H{r}-G{r})/G{r})'
        ws[f"J{r}"].number_format = '+0.00%;-0.00%;0.00%'
        ws[f"J{r}"].fill = fill(C_LOCKED_BG)
        ws[f"J{r}"].alignment = right()
        ws[f"J{r}"].border = border_all()

    n = len(costs)
    sr = 5 + n + 1
    ws[f"B{sr}"].value = "TOTAL"
    ws[f"B{sr}"].font = Font(bold=True, size=11, name="Calibri", color=C_WHITE)
    ws[f"B{sr}"].fill = fill(C_NAVY)
    ws[f"B{sr}"].alignment = center()
    ws[f"B{sr}"].border = border_all()
    ws.row_dimensions[sr].height = 22

    for col in "CDEFIJ":
        ws[f"{col}{sr}"].fill = fill(C_NAVY)
        ws[f"{col}{sr}"].border = border_all()

    for col, fmt in [("G","€#,##0.00"),("H","€#,##0.00"),("I",'+€#,##0.00;-€#,##0.00;€0.00')]:
        ws[f"{col}{sr}"].value = f"=SUM({col}5:{col}{sr-1})"
        ws[f"{col}{sr}"].number_format = fmt
        ws[f"{col}{sr}"].font = Font(bold=True, size=11, name="Calibri", color=C_WHITE)
        ws[f"{col}{sr}"].fill = fill(C_NAVY)
        ws[f"{col}{sr}"].alignment = right()
        ws[f"{col}{sr}"].border = border_all()

    ws[f"J{sr}"].value = f'=IF(G{sr}=0,"N/A",(H{sr}-G{sr})/G{sr})'
    ws[f"J{sr}"].number_format = '+0.00%;-0.00%;0.00%'
    ws[f"J{sr}"].font = Font(bold=True, size=11, name="Calibri", color=C_WHITE)
    ws[f"J{sr}"].fill = fill(C_NAVY)
    ws[f"J{sr}"].alignment = right()
    ws[f"J{sr}"].border = border_all()

    ws.conditional_formatting.add(
        f"I5:I{sr-1}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=fill(C_INCREASE))
    )
    ws.conditional_formatting.add(
        f"I5:I{sr-1}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(C_DECREASE))
    )

# ─── Sheet: Export Preview ────────────────────────────────────────────────────
def build_export_preview(wb, today, year, q):
    ws = wb.create_sheet("📤 Export Preview")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDE", [4, 80, 20, 20, 20]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:E2")
    ws["B2"].value = "📤  Export Preview — After filling in all tabs, run the import script"
    ws["B2"].font = hfont(size=13)
    ws["B2"].fill = fill(C_NAVY)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 32

    instructions = [
        "",
        "HOW TO APPLY YOUR CHANGES TO THE DATABASE:",
        "",
        "  1. Save this workbook.",
        f"  2. Open a terminal in the repository root.",
        f"  3. Run:   python tools/import_from_quarterly_excel.py",
        f"     (optionally:  --file data/quarterly_update_{quarter_label(year,q)}.xlsx)",
        "  4. The script will:",
        "       • Read all yellow 'New ...' cells from this workbook",
        "       • Skip blank cells (no change applied)",
        "       • Write updated rows back to the CSV files in data/",
        "       • Append a history entry to data/history.csv with a timestamp",
        "       • Print a summary of every field changed",
        "  5. Review the diff:   git diff data/",
        "  6. Commit and push:   git add data/ && git commit -m 'Quarterly cost update <quarter>'",
        "",
        "NEXT QUARTERLY REVIEW:",
        f"  Due date: {(today + timedelta(days=91)).strftime('%d %B %Y')}",
        f"  Regenerate this workbook by running:",
        "    python tools/generate_quarterly_update_excel.py",
    ]
    for i, line in enumerate(instructions, start=3):
        ws.row_dimensions[i + 2].height = 18
        ws[f"B{i+2}"].value = line
        ws[f"B{i+2}"].font = Font(bold=(":" in line and not line.startswith(" ")),
                                   size=10, name="Calibri",
                                   color=C_NAVY if not line.startswith("  ") else "000000")
        ws[f"B{i+2}"].alignment = left()
        ws.merge_cells(f"B{i+2}:E{i+2}")

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    today = date.today()
    year, q = current_quarter(today)
    qlabel = quarter_label(year, q)
    out_path = DATA_DIR / f"quarterly_update_{qlabel}.xlsx"

    print(f"🏗  Building quarterly update workbook for {qlabel} …")

    wb = Workbook()
    del wb[wb.sheetnames[0]]

    build_instructions(wb, today, year, q)
    build_schedule(wb, today, year, q)
    build_materials(wb)
    build_processes(wb)
    build_supplier_quotes(wb, today)
    build_market(wb)
    build_cost_impact(wb)
    build_export_preview(wb, today, year, q)

    wb.active = wb["📋 Instructions"]

    wb.save(out_path)
    print(f"✅  Saved:  {out_path.relative_to(REPO_ROOT)}")
    print()
    print("Next steps:")
    print("  1. Open the workbook and fill in the yellow cells.")
    print("  2. Run:  python tools/import_from_quarterly_excel.py")
    print("  3. Commit the updated data/ folder.")


if __name__ == "__main__":
    main()
