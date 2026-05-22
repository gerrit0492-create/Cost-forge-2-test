"""
Quarterly Cost Update — Import from Excel
==========================================
Reads the filled-in quarterly update workbook and writes the changes
back to the CSV database files in data/.

Usage:
    python tools/import_from_quarterly_excel.py
    python tools/import_from_quarterly_excel.py --file data/quarterly_update_2026-Q2.xlsx
    python tools/import_from_quarterly_excel.py --dry-run   # preview only, no files written

Requirements:
    pip install openpyxl
"""

import argparse
import csv
import os
import sys
import glob
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl not found. Run:  pip install openpyxl")

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR  = REPO_ROOT / "data"

# ─── CLI ──────────────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(description="Import quarterly Excel back to CSV databases.")
    p.add_argument("--file",    default=None,  help="Path to the .xlsx workbook (auto-detected if omitted).")
    p.add_argument("--dry-run", action="store_true", help="Print changes only; do not write files.")
    return p.parse_args()

# ─── CSV helpers ─────────────────────────────────────────────────────────────
def load_csv(path):
    rows = []
    if not path.exists():
        return [], []
    with open(path, newline="", encoding="utf-8") as f:
        raw = f.read()
    # Strip comment lines
    lines = [l for l in raw.splitlines() if not l.strip().startswith("#")]
    reader = csv.DictReader(lines)
    headers = list(reader.fieldnames or [])
    for row in reader:
        rows.append(dict(row))
    return headers, rows

def save_csv(path, headers, rows, dry_run=False):
    if dry_run:
        print(f"  [dry-run] Would write {len(rows)} rows → {path.relative_to(REPO_ROOT)}")
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  ✅  Written {len(rows)} rows → {path.relative_to(REPO_ROOT)}")

def get_cell(ws, col, row):
    """Return cell value; strip whitespace from strings."""
    val = ws[f"{col}{row}"].value
    if isinstance(val, str):
        val = val.strip()
    return val

def is_blank(val):
    return val is None or str(val).strip() == ""

def safe_float(val):
    try:
        return float(str(val).replace(",", "."))
    except Exception:
        return None

# ─── Materials import ─────────────────────────────────────────────────────────
def import_materials(wb, dry_run):
    print("\n── 💎 Materials DB ──────────────────────────────────────────")
    ws = wb["💎 Materials DB"]
    headers, rows = load_csv(DATA_DIR / "materials_db.csv")
    if not rows:
        print("  ⚠️  materials_db.csv not found or empty — skipping.")
        return 0

    changes = 0
    for idx, row in enumerate(rows):
        excel_row = idx + 5           # data starts at row 5
        new_price = get_cell(ws, "F", excel_row)
        if is_blank(new_price):
            continue
        nf = safe_float(new_price)
        if nf is None:
            print(f"  ⚠️  Row {excel_row}: invalid price '{new_price}' for {row['material_id']} — skipped.")
            continue
        old_val = row.get("price_eur_per_kg","")
        try:
            pct = ((nf - float(old_val)) / float(old_val) * 100) if old_val else 0
        except Exception:
            pct = 0
        print(f"  {row['material_id']:12s}  {old_val:>8} → {nf:>8.4f} €/kg  ({pct:+.1f}%)")
        rows[idx]["price_eur_per_kg"] = f"{nf:.4f}"
        changes += 1

    if changes == 0:
        print("  ℹ️  No material prices updated (all New Price cells blank).")
    else:
        save_csv(DATA_DIR / "materials_db.csv", headers, rows, dry_run)
    return changes

# ─── Process Rates import ─────────────────────────────────────────────────────
def import_processes(wb, dry_run):
    print("\n── ⚙️ Process Rates ─────────────────────────────────────────")
    ws = wb["⚙️ Process Rates"]
    headers, rows = load_csv(DATA_DIR / "processes_db.csv")
    if not rows:
        print("  ⚠️  processes_db.csv not found or empty — skipping.")
        return 0

    changes = 0
    for idx, row in enumerate(rows):
        excel_row = idx + 5
        changed = False

        field_map = [
            ("E", "machine_rate_eur_h",  "Machine rate"),
            ("G", "labor_rate_eur_h",    "Labour rate"),
            ("I", "overhead_pct",        "Overhead %"),
            ("K", "margin_pct",          "Margin %"),
        ]
        for col, field, label in field_map:
            new_val = get_cell(ws, col, excel_row)
            if is_blank(new_val):
                continue
            nf = safe_float(new_val)
            if nf is None:
                print(f"  ⚠️  {row['process_id']} {label}: invalid '{new_val}' — skipped.")
                continue
            old = row.get(field,"")
            print(f"  {row['process_id']:18s}  {label:16s}: {old:>8} → {nf:.4f}")
            rows[idx][field] = f"{nf:.4f}"
            changed = True

        if changed:
            changes += 1

    if changes == 0:
        print("  ℹ️  No process rates updated.")
    else:
        save_csv(DATA_DIR / "processes_db.csv", headers, rows, dry_run)
    return changes

# ─── Supplier Quotes import ───────────────────────────────────────────────────
def import_supplier_quotes(wb, dry_run):
    print("\n── 🏢 Supplier Quotes ───────────────────────────────────────")
    ws = wb["🏢 Supplier Quotes"]

    _, sq_rows = load_csv(DATA_DIR / "supplier_quotes.csv")
    _, q_rows  = load_csv(DATA_DIR / "quotes.csv")

    sq_headers = ["supplier","material_id","price_eur_per_kg","lead_time_days","valid_until","preferred"]
    q_headers  = ["supplier","material_id","price_eur_per_kg","lead_time_days","valid_until","preferred"]

    # Build unified list matching how the generator combined them
    all_rows   = sq_rows + [r for r in q_rows
                             if not any(s["material_id"]==r["material_id"] and
                                        s["supplier"]==r["supplier"] for s in sq_rows)]

    # Build index maps: (supplier, material_id) → position in sq_rows / q_rows
    sq_key = {(r["supplier"], r["material_id"]): i for i, r in enumerate(sq_rows)}
    q_key  = {(r["supplier"], r["material_id"]): i for i, r in enumerate(q_rows)}

    sq_changes = 0
    q_changes  = 0

    for idx, row in enumerate(all_rows):
        excel_row = idx + 5
        supplier  = row["supplier"]
        mat_id    = row["material_id"]
        key       = (supplier, mat_id)

        # E = new price
        new_price = get_cell(ws, "E", excel_row)
        nf = safe_float(new_price) if not is_blank(new_price) else None

        # H = new lead time
        new_lead = get_cell(ws, "H", excel_row)
        nl = safe_float(new_lead) if not is_blank(new_lead) else None

        # J = new valid_until
        new_valid = get_cell(ws, "J", excel_row)
        nv = str(new_valid).strip() if not is_blank(new_valid) else None
        if isinstance(new_valid, datetime):
            nv = new_valid.strftime("%Y-%m-%d")
        elif isinstance(new_valid, date):
            nv = new_valid.strftime("%Y-%m-%d")

        # K = preferred
        new_pref = get_cell(ws, "K", excel_row)
        np_ = str(int(safe_float(new_pref))) if not is_blank(new_pref) else None

        if nf is None and nl is None and nv is None and np_ is None:
            continue

        summary_parts = []
        if nf is not None:
            old_p = row.get("price_eur_per_kg","")
            summary_parts.append(f"price {old_p}→{nf:.4f}")
        if nl is not None:
            summary_parts.append(f"lead {row.get('lead_time_days','')}→{int(nl)}")
        if nv is not None:
            summary_parts.append(f"valid_until {row.get('valid_until','')}→{nv}")
        if np_ is not None:
            summary_parts.append(f"preferred {row.get('preferred','')}→{np_}")

        print(f"  {supplier:24s} {mat_id:14s}  {', '.join(summary_parts)}")

        if key in sq_key:
            i = sq_key[key]
            if nf  is not None: sq_rows[i]["price_eur_per_kg"] = f"{nf:.4f}"
            if nl  is not None: sq_rows[i]["lead_time_days"]   = str(int(nl))
            if nv  is not None: sq_rows[i]["valid_until"]      = nv
            if np_ is not None: sq_rows[i]["preferred"]        = np_
            sq_changes += 1
        elif key in q_key:
            i = q_key[key]
            if nf  is not None: q_rows[i]["price_eur_per_kg"] = f"{nf:.4f}"
            if nl  is not None: q_rows[i]["lead_time_days"]   = str(int(nl))
            if nv  is not None: q_rows[i]["valid_until"]      = nv
            if np_ is not None: q_rows[i]["preferred"]        = np_
            q_changes += 1

    if sq_changes == 0 and q_changes == 0:
        print("  ℹ️  No supplier quote changes.")
    else:
        if sq_changes > 0:
            save_csv(DATA_DIR / "supplier_quotes.csv", sq_headers, sq_rows, dry_run)
        if q_changes > 0:
            save_csv(DATA_DIR / "quotes.csv", q_headers, q_rows, dry_run)

    return sq_changes + q_changes

# ─── Market Adjustments import ────────────────────────────────────────────────
def import_market_factors(wb, dry_run):
    print("\n── 📊 Market Adjustments ────────────────────────────────────")
    ws = wb["📊 Market Adjustments"]
    headers, rows = load_csv(DATA_DIR / "market-factors.csv")
    if not rows:
        print("  ⚠️  market-factors.csv not found or empty — skipping.")
        return 0

    changes = 0
    for idx, row in enumerate(rows):
        excel_row = idx + 5
        new_pct    = get_cell(ws, "E", excel_row)
        new_factor = get_cell(ws, "G", excel_row)
        any_change = False

        if not is_blank(new_pct):
            nf = safe_float(new_pct)
            if nf is not None:
                print(f"  {str(row.get('material_id','(all)') or '(all)'):12s}  "
                      f"pct_change: {row.get('pct_change','')} → {nf}")
                rows[idx]["pct_change"] = f"{nf}"
                rows[idx]["factor"]     = ""
                any_change = True

        if not is_blank(new_factor):
            nf = safe_float(new_factor)
            if nf is not None:
                print(f"  {str(row.get('material_id','(all)') or '(all)'):12s}  "
                      f"factor: {row.get('factor','')} → {nf}")
                rows[idx]["factor"]     = f"{nf}"
                rows[idx]["pct_change"] = ""
                any_change = True

        if any_change:
            changes += 1

    if changes == 0:
        print("  ℹ️  No market factor changes.")
    else:
        out_path = DATA_DIR / "market-factors.csv"
        if dry_run:
            print(f"  [dry-run] Would write {len(rows)} rows → {out_path.relative_to(REPO_ROOT)}")
        else:
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                f.write("# Marktfactoren — match op material_id of commodity. Vul pct_change (%) of factor in.\n")
                writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows)
            print(f"  ✅  Written {len(rows)} rows → {out_path.relative_to(REPO_ROOT)}")
    return changes

# ─── History log ─────────────────────────────────────────────────────────────
def append_history_entry(dry_run, totals: dict, source_file: str):
    hist_path = DATA_DIR / "log.csv"
    today_str = date.today().isoformat()
    entry = {
        "date":     today_str,
        "action":   "quarterly_update",
        "source":   Path(source_file).name,
        "materials_changed": totals.get("materials", 0),
        "processes_changed": totals.get("processes", 0),
        "quotes_changed":    totals.get("quotes", 0),
        "market_changed":    totals.get("market", 0),
        "note":              f"Quarterly cost update imported from {Path(source_file).name}",
    }
    if dry_run:
        print(f"\n  [dry-run] Would append to {hist_path.relative_to(REPO_ROOT)}: {entry}")
        return
    exists = hist_path.exists()
    with open(hist_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(entry.keys()))
        if not exists:
            writer.writeheader()
        writer.writerow(entry)
    print(f"\n  ✅  History entry appended to {hist_path.relative_to(REPO_ROOT)}")

# ─── Auto-detect workbook ─────────────────────────────────────────────────────
def find_latest_workbook():
    pattern = str(DATA_DIR / "quarterly_update_*.xlsx")
    files   = sorted(glob.glob(pattern), reverse=True)
    if not files:
        return None
    return files[0]

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()

    if args.file:
        wb_path = Path(args.file)
        if not wb_path.is_absolute():
            wb_path = REPO_ROOT / wb_path
    else:
        wb_path = find_latest_workbook()
        if not wb_path:
            sys.exit("❌  No quarterly_update_*.xlsx found in data/. "
                     "Run generate_quarterly_update_excel.py first.")

    if not Path(wb_path).exists():
        sys.exit(f"❌  File not found: {wb_path}")

    print(f"\n{'='*62}")
    print(f"  Cost Forge — Quarterly Cost Import")
    print(f"  Workbook : {Path(wb_path).name}")
    if args.dry_run:
        print(f"  Mode     : DRY RUN (no files will be written)")
    print(f"{'='*62}")

    wb = load_workbook(str(wb_path), data_only=True)

    required = ["💎 Materials DB","⚙️ Process Rates","🏢 Supplier Quotes","📊 Market Adjustments"]
    missing  = [s for s in required if s not in wb.sheetnames]
    if missing:
        sys.exit(f"❌  Missing sheets: {missing}\n"
                 "    Make sure you are using the workbook generated by generate_quarterly_update_excel.py")

    totals = {}
    totals["materials"] = import_materials(wb, args.dry_run)
    totals["processes"] = import_processes(wb, args.dry_run)
    totals["quotes"]    = import_supplier_quotes(wb, args.dry_run)
    totals["market"]    = import_market_factors(wb, args.dry_run)

    append_history_entry(args.dry_run, totals, str(wb_path))

    total_changes = sum(totals.values())
    print(f"\n{'='*62}")
    print(f"  Summary")
    print(f"  {'Materials changed:':<28} {totals['materials']}")
    print(f"  {'Process rows changed:':<28} {totals['processes']}")
    print(f"  {'Supplier quote rows changed:':<28} {totals['quotes']}")
    print(f"  {'Market factor rows changed:':<28} {totals['market']}")
    print(f"  {'─'*30}")
    print(f"  {'TOTAL changes applied:':<28} {total_changes}")
    if args.dry_run:
        print("\n  ⚠️  DRY RUN — no files were written.")
    else:
        print("\n  Next step:  git diff data/   (review changes)")
        print("              git add data/ && git commit -m 'Quarterly cost update'")
    print(f"{'='*62}\n")


if __name__ == "__main__":
    main()
