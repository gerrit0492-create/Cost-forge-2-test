"""
Quarterly Cost Update — Streamlit page
=======================================
Generate the quarterly update workbook, or import a filled one back
into the CSV databases without touching the command line.
"""
from __future__ import annotations

import io
import csv
import glob
import logging
from datetime import date, timedelta
from pathlib import Path

import streamlit as st

from utils.nav import home_button
from utils.safe import guard

logger = logging.getLogger(__name__)

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR  = REPO_ROOT / "data"


# ─── Quarter helpers ──────────────────────────────────────────────────────────
def current_quarter(d: date = None):
    d = d or date.today()
    q = (d.month - 1) // 3 + 1
    return d.year, q

def quarter_label(year, q):
    return f"{year}-Q{q}"

def next_quarter_start(d: date = None):
    d = d or date.today()
    year, q = current_quarter(d)
    if q == 4:
        return date(year + 1, 1, 1)
    return date(year, q * 3 + 1, 1)


# ─── Generate Excel (calls the existing generator) ────────────────────────────
def generate_excel_bytes() -> bytes:
    """Run the generator in-memory and return xlsx bytes."""
    try:
        from openpyxl import Workbook
    except ImportError:
        st.error("openpyxl is not installed. Run: pip install openpyxl")
        st.stop()

    gen_path = REPO_ROOT / "tools" / "generate_quarterly_update_excel.py"
    if not gen_path.exists():
        raise FileNotFoundError(f"Generator not found: {gen_path}")

    # __file__ must be seeded so the generator can resolve REPO_ROOT via Path(__file__)
    ns: dict = {"__file__": str(gen_path)}
    exec(compile(gen_path.read_text(), str(gen_path), "exec"), ns)

    today = date.today()
    year, q = ns["current_quarter"](today)

    wb = Workbook()
    del wb[wb.sheetnames[0]]

    ns["build_instructions"](wb, today, year, q)
    ns["build_schedule"](wb, today, year, q)
    ns["build_materials"](wb)
    ns["build_processes"](wb)
    ns["build_supplier_quotes"](wb, today)
    ns["build_market"](wb)
    ns["build_cost_impact"](wb)
    ns["build_export_preview"](wb, today, year, q)

    wb.active = wb["📋 Instructions"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Import helpers (inline; mirrors import_from_quarterly_excel.py) ──────────
def load_csv_db(path: Path):
    if not path.exists():
        return [], []
    with open(path, newline="", encoding="utf-8") as f:
        raw = f.read()
    lines = [l for l in raw.splitlines() if not l.strip().startswith("#")]
    reader = csv.DictReader(lines)
    headers = list(reader.fieldnames or [])
    rows = [dict(r) for r in reader]
    return headers, rows

def save_csv_db(path: Path, headers, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

def is_blank(val):
    return val is None or str(val).strip() == ""

def safe_float(val):
    try:
        return float(str(val).replace(",", "."))
    except Exception:
        return None

def run_import(wb, dry_run: bool) -> dict:
    """
    Import from a filled quarterly workbook.
    Returns a summary dict: {sheet: [(field, old, new), ...]}
    """
    from datetime import datetime as DT
    changes: dict[str, list] = {
        "Materials": [], "Processes": [], "Supplier Quotes": [], "Market Adjustments": []
    }

    # ── Materials ──
    sheet_name = "💎 Materials DB"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers, rows = load_csv_db(DATA_DIR / "materials_db.csv")
        for idx, row in enumerate(rows):
            r = idx + 5
            new_price = ws[f"F{r}"].value
            if is_blank(new_price):
                continue
            nf = safe_float(new_price)
            if nf is None:
                continue
            old = row.get("price_eur_per_kg", "")
            changes["Materials"].append((row["material_id"], "price_eur_per_kg", old, f"{nf:.4f}"))
            rows[idx]["price_eur_per_kg"] = f"{nf:.4f}"
        if not dry_run and changes["Materials"]:
            save_csv_db(DATA_DIR / "materials_db.csv", headers, rows)

    # ── Processes ──
    sheet_name = "⚙️ Process Rates"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers, rows = load_csv_db(DATA_DIR / "processes_db.csv")
        field_map = [("E","machine_rate_eur_h"),("G","labor_rate_eur_h"),
                     ("I","overhead_pct"),("K","margin_pct")]
        for idx, row in enumerate(rows):
            r = idx + 5
            for col, field in field_map:
                val = ws[f"{col}{r}"].value
                if is_blank(val):
                    continue
                nf = safe_float(val)
                if nf is None:
                    continue
                old = row.get(field, "")
                changes["Processes"].append((row["process_id"], field, old, f"{nf:.4f}"))
                rows[idx][field] = f"{nf:.4f}"
        if not dry_run and changes["Processes"]:
            save_csv_db(DATA_DIR / "processes_db.csv", headers, rows)

    # ── Supplier Quotes ──
    sheet_name = "🏢 Supplier Quotes"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _, sq_rows = load_csv_db(DATA_DIR / "supplier_quotes.csv")
        _, q_rows  = load_csv_db(DATA_DIR / "quotes.csv")
        sq_headers = ["supplier","material_id","price_eur_per_kg","lead_time_days","valid_until","preferred"]
        q_headers  = sq_headers[:]
        all_rows   = sq_rows + [r for r in q_rows if not any(
            s["material_id"]==r["material_id"] and s["supplier"]==r["supplier"] for s in sq_rows)]
        sq_key = {(r["supplier"],r["material_id"]): i for i,r in enumerate(sq_rows)}
        q_key  = {(r["supplier"],r["material_id"]): i for i,r in enumerate(q_rows)}
        sq_changed, q_changed = False, False
        for idx, row in enumerate(all_rows):
            r = idx + 5
            key = (row["supplier"], row["material_id"])
            new_price = ws[f"E{r}"].value
            new_lead  = ws[f"H{r}"].value
            new_valid = ws[f"J{r}"].value
            new_pref  = ws[f"K{r}"].value

            nf = safe_float(new_price) if not is_blank(new_price) else None
            nl = safe_float(new_lead)  if not is_blank(new_lead)  else None
            nv = None
            if not is_blank(new_valid):
                if isinstance(new_valid, (date,)):
                    nv = new_valid.strftime("%Y-%m-%d")
                elif isinstance(new_valid, DT):
                    nv = new_valid.strftime("%Y-%m-%d")
                else:
                    nv = str(new_valid).strip()
            np_ = str(int(safe_float(new_pref))) if not is_blank(new_pref) and safe_float(new_pref) is not None else None

            if nf is None and nl is None and nv is None and np_ is None:
                continue

            parts = []
            if nf is not None:
                parts.append(f"price {row.get('price_eur_per_kg','')} → {nf:.4f}")
            if nl is not None:
                parts.append(f"lead {row.get('lead_time_days','')} → {int(nl)}")
            if nv is not None:
                parts.append(f"valid_until → {nv}")
            changes["Supplier Quotes"].append((f"{row['supplier']} / {row['material_id']}", ", ".join(parts), "", ""))

            if key in sq_key:
                i = sq_key[key]
                if nf  is not None: sq_rows[i]["price_eur_per_kg"] = f"{nf:.4f}"
                if nl  is not None: sq_rows[i]["lead_time_days"]   = str(int(nl))
                if nv  is not None: sq_rows[i]["valid_until"]      = nv
                if np_ is not None: sq_rows[i]["preferred"]        = np_
                sq_changed = True
            elif key in q_key:
                i = q_key[key]
                if nf  is not None: q_rows[i]["price_eur_per_kg"] = f"{nf:.4f}"
                if nl  is not None: q_rows[i]["lead_time_days"]   = str(int(nl))
                if nv  is not None: q_rows[i]["valid_until"]      = nv
                if np_ is not None: q_rows[i]["preferred"]        = np_
                q_changed = True

        if not dry_run:
            if sq_changed:
                save_csv_db(DATA_DIR / "supplier_quotes.csv", sq_headers, sq_rows)
            if q_changed:
                save_csv_db(DATA_DIR / "quotes.csv", q_headers, q_rows)

    # ── Market Adjustments ──
    sheet_name = "📊 Market Adjustments"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers, rows = load_csv_db(DATA_DIR / "market-factors.csv")
        for idx, row in enumerate(rows):
            r = idx + 5
            new_pct    = ws[f"E{r}"].value
            new_factor = ws[f"G{r}"].value
            changed = False
            if not is_blank(new_pct):
                nf = safe_float(new_pct)
                if nf is not None:
                    old = row.get("pct_change","")
                    changes["Market Adjustments"].append(
                        (str(row.get("material_id","(all)")), "pct_change", old, str(nf)))
                    rows[idx]["pct_change"] = str(nf)
                    rows[idx]["factor"] = ""
                    changed = True
            if not is_blank(new_factor):
                nf = safe_float(new_factor)
                if nf is not None:
                    old = row.get("factor","")
                    changes["Market Adjustments"].append(
                        (str(row.get("material_id","(all)")), "factor", old, str(nf)))
                    rows[idx]["factor"] = str(nf)
                    rows[idx]["pct_change"] = ""
                    changed = True
        if not dry_run and changes["Market Adjustments"]:
            out_path = DATA_DIR / "market-factors.csv"
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                f.write("# Marktfactoren — match op material_id of commodity.\n")
                writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows)

    # ── History entry ──
    if not dry_run:
        total = sum(len(v) for v in changes.values())
        if total > 0:
            hist_path = DATA_DIR / "log.csv"
            entry = {
                "date":   date.today().isoformat(),
                "action": "quarterly_update",
                "source": "Streamlit upload",
                "materials_changed": len(changes["Materials"]),
                "processes_changed": len(changes["Processes"]),
                "quotes_changed":    len(changes["Supplier Quotes"]),
                "market_changed":    len(changes["Market Adjustments"]),
                "note":  "Imported via Quarterly Update page",
            }
            exists = hist_path.exists()
            with open(hist_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=list(entry.keys()))
                if not exists:
                    writer.writeheader()
                writer.writerow(entry)

    return changes


# ─── UI ───────────────────────────────────────────────────────────────────────
def main():
    home_button()

    today = date.today()
    year, q = current_quarter(today)
    qlabel = quarter_label(year, q)
    next_start = next_quarter_start(today)
    days_left = (next_start - today).days

    st.title("🔄 Quarterly Cost Update")
    st.caption(f"Current quarter: **{qlabel}** · Next review in **{days_left} days** ({next_start.strftime('%d %b %Y')})")

    # ── Progress bar ──────────────────────────────────────────────────────────
    pct = max(0, min(1, 1 - days_left / 91))
    st.progress(pct, text=f"Quarter progress: {pct*100:.0f}%")

    st.divider()

    tab_gen, tab_import, tab_history = st.tabs([
        "📥 1 — Generate Workbook",
        "📤 2 — Import Filled Workbook",
        "📜 3 — History",
    ])

    # ── TAB 1: Generate ───────────────────────────────────────────────────────
    with tab_gen:
        st.subheader("Generate quarterly update workbook")
        st.markdown(
            "Produces a pre-loaded Excel workbook with your current database values. "
            "Fill in the **yellow columns**, then upload it in Tab 2 to apply the changes."
        )

        col_a, col_b = st.columns([3, 1])
        with col_a:
            st.info(
                "The workbook includes:\n"
                "- 💎 Materials DB — current prices + empty 'New Price' column\n"
                "- ⚙️ Process Rates — machine / labour / overhead\n"
                "- 🏢 Supplier Quotes — expiry highlighted red/amber\n"
                "- 📊 Market Adjustments — commodity factors\n"
                "- 💰 Cost Impact — live formula preview"
            )

        if col_b.button("⚙️ Build workbook", use_container_width=True):
            with st.spinner("Building workbook from current database …"):
                try:
                    xlsx_bytes = generate_excel_bytes()
                    st.session_state["gen_xlsx"] = xlsx_bytes
                    st.session_state["gen_label"] = qlabel
                    st.success(f"✅ Workbook ready — {len(xlsx_bytes)//1024} KB")
                except Exception as e:
                    import traceback
                    st.error(f"Generation failed: {e}")
                    st.code(traceback.format_exc(), language="python")
                    logger.exception("generate_excel_bytes failed")

        if "gen_xlsx" in st.session_state:
            st.download_button(
                label=f"⬇️ Download quarterly_update_{st.session_state['gen_label']}.xlsx",
                data=st.session_state["gen_xlsx"],
                file_name=f"quarterly_update_{st.session_state['gen_label']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.divider()
        st.subheader("Or use the pre-generated file")
        st.caption("Any `quarterly_update_*.xlsx` already in `data/` can be downloaded here.")

        existing = sorted(glob.glob(str(DATA_DIR / "quarterly_update_*.xlsx")), reverse=True)
        if existing:
            for path in existing:
                fname = Path(path).name
                with open(path, "rb") as f:
                    st.download_button(
                        f"⬇️ {fname}",
                        data=f.read(),
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{fname}",
                    )
        else:
            st.caption("No pre-generated files found in data/")

    # ── TAB 2: Import ─────────────────────────────────────────────────────────
    with tab_import:
        st.subheader("Import a filled workbook back to the database")
        st.markdown(
            "Upload the workbook you filled in. Only **non-blank yellow cells** are applied — "
            "blank cells are skipped (no change)."
        )

        uploaded = st.file_uploader(
            "Upload quarterly_update_*.xlsx",
            type=["xlsx"],
            help="Must be a workbook generated by Cost Forge (correct sheet names required).",
        )

        dry_run = st.toggle("🔍 Dry run (preview only — no files written)", value=True)

        if dry_run:
            st.info("Dry run is ON — review changes below before applying.")

        if uploaded and st.button("▶️ Run import", type="primary", use_container_width=True):
            try:
                from openpyxl import load_workbook as _lw
                wb = _lw(io.BytesIO(uploaded.read()), data_only=True)

                required = ["💎 Materials DB", "⚙️ Process Rates",
                            "🏢 Supplier Quotes", "📊 Market Adjustments"]
                missing = [s for s in required if s not in wb.sheetnames]
                if missing:
                    st.error(f"Wrong file format — missing sheets: {missing}")
                    st.stop()

                with st.spinner("Importing …"):
                    changes = run_import(wb, dry_run=dry_run)

                total = sum(len(v) for v in changes.values())

                if total == 0:
                    st.warning("No filled-in cells found — nothing to import. "
                               "Make sure you filled in the yellow columns.")
                else:
                    if dry_run:
                        st.success(f"**Dry run complete** — {total} change(s) detected. "
                                   "Disable dry run and re-run to apply.")
                    else:
                        st.success(f"✅ **{total} change(s) applied to the database.**")

                    for section, rows in changes.items():
                        if not rows:
                            continue
                        st.markdown(f"**{section}** — {len(rows)} change(s)")
                        if section == "Supplier Quotes":
                            for label, summary, _, __ in rows:
                                st.markdown(f"  - `{label}`: {summary}")
                        else:
                            for item_id, field, old, new in rows:
                                arrow = "🔺" if (safe_float(new) or 0) > (safe_float(old) or 0) else "🔻"
                                st.markdown(f"  - `{item_id}` · {field}: "
                                            f"**{old}** → **{new}** {arrow}")

            except Exception as e:
                st.error(f"Import failed: {e}")
                logger.exception("run_import failed")

    # ── TAB 3: History ────────────────────────────────────────────────────────
    with tab_history:
        st.subheader("Update history log")
        log_path = DATA_DIR / "log.csv"
        if not log_path.exists():
            st.info("No log entries yet. Log file (`data/log.csv`) will be created after the first import.")
        else:
            try:
                import pandas as pd
                df = pd.read_csv(log_path)
                if df.empty:
                    st.info("Log file exists but has no entries yet.")
                else:
                    if "action" in df.columns:
                        q_df = df[df["action"] == "quarterly_update"].copy()
                    else:
                        q_df = df.copy()

                    if q_df.empty:
                        st.info("No quarterly update entries in the log yet.")
                    else:
                        st.dataframe(
                            q_df.sort_values("date", ascending=False).reset_index(drop=True),
                            use_container_width=True,
                        )
            except Exception as e:
                st.error(f"Could not read log: {e}")

        st.divider()
        st.subheader("📅 Next quarterly review")
        next_q_year, next_q = current_quarter(next_start)
        st.metric(
            label=f"Quarter {quarter_label(next_q_year, next_q)}",
            value=next_start.strftime("%d %B %Y"),
            delta=f"in {days_left} days",
        )
        st.caption(
            "Recommended: run **Generate Workbook** (Tab 1) 2 weeks before the due date "
            "and share it with your purchasing team."
        )


guard(main)
