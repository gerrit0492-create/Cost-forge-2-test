"""QMS Component Price Database
==========================================
Central price database for waterjet QMS components across supply chains
(IN01 = India, NL07 = EU/Netherlands) and all waterjet sizes.

Features
--------
- Live KPI cards per supply-chain × size combination
- Filterable price list with BOM-total metric
- In-place data-editor for adding / editing items
- Surcharge / price adjustment tool with preview
- Excel workbook generation (one sheet per supply chain)
- Excel import with dry-run and change diff
- Full audit log
"""
from __future__ import annotations

import io
import csv
import logging
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st

from utils.nav import home_button
from utils.safe import guard

logger = logging.getLogger(__name__)

REPO_ROOT     = Path(__file__).resolve().parent.parent
DATA_DIR      = REPO_ROOT / "data"
QMS_CSV       = DATA_DIR / "qms_prices.csv"
QMS_LOG       = DATA_DIR / "qms_log.csv"
SUPPLY_CHAINS = ["IN01", "NL07"]
CATEGORIES    = ["Hydraulics", "Sealing", "Structural", "Drive", "Steering", "Electrical", "Other"]

SC_LABELS = {"IN01": "\U0001f1ee\U0001f1f3 IN01 (India)", "NL07": "\U0001f1f3\U0001f1f1 NL07 (EU)"}

# ─── Colour palette (matches generate_quarterly_update_excel.py) ───────────────────────────
C_NAVY      = "0D3B66"
C_BLUE      = "1565C0"
C_TEAL      = "00838F"
C_INPUT_BG  = "FFFDE7"   # pale yellow  – editable cells
C_LOCKED_BG = "ECEFF1"   # pale grey    – formula / locked
C_WHITE     = "FFFFFF"
C_STRIPE    = "EBF3FF"
C_LIGHT_BG  = "F5F7FA"
C_INCREASE  = "FFCDD2"
C_DECREASE  = "C8E6C9"


# ─── Data I/O ────────────────────────────────────────────────────────────────────────────────

def load_qms() -> pd.DataFrame:
    """Read qms_prices.csv and return a DataFrame with a computed line_total_eur column."""
    if not QMS_CSV.exists():
        return pd.DataFrame(columns=[
            "item_id", "description", "category", "supply_chain",
            "size_mm", "qty", "price_eur", "surcharge_pct",
            "supplier", "valid_from", "notes",
        ])
    df = pd.read_csv(QMS_CSV, dtype=str)
    for col in ("price_eur", "surcharge_pct", "qty"):
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    df["size_mm"] = pd.to_numeric(df["size_mm"], errors="coerce").fillna(0).astype(int)
    df["line_total_eur"] = df["qty"] * df["price_eur"] * (1 + df["surcharge_pct"] / 100)
    return df


def save_qms(df: pd.DataFrame) -> None:
    """Write DataFrame back to qms_prices.csv (drops the computed column)."""
    out = df.drop(columns=["line_total_eur"], errors="ignore")
    out.to_csv(QMS_CSV, index=False)


def _log_update(action: str, detail: str, n_changes: int) -> None:
    """Append one row to qms_log.csv."""
    entry = {
        "date":      date.today().isoformat(),
        "action":    action,
        "detail":    detail,
        "n_changes": n_changes,
    }
    exists = QMS_LOG.exists()
    with open(QMS_LOG, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(entry.keys()))
        if not exists:
            writer.writeheader()
        writer.writerow(entry)


# ─── Excel helpers ─────────────────────────────────────────────────────────────────────────────

def _build_qms_excel(df: pd.DataFrame) -> bytes:
    """Build a QMS price update workbook (one sheet per supply chain) and return bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        st.error("openpyxl is not installed — run: pip install openpyxl")
        st.stop()

    today_str = date.today().strftime("%d %b %Y")

    # ── Style helpers (inline so function is self-contained) ──────────────────────────────
    def _fill(hex_c: str):
        return PatternFill("solid", fgColor=hex_c)

    def _font(bold=False, size=10, color="000000"):
        return Font(bold=bold, size=size, color=color, name="Calibri")

    def _hfont(size=10):
        return Font(bold=True, size=size, color=C_WHITE, name="Calibri")

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _right():
        return Alignment(horizontal="right", vertical="center")

    _thin  = Side(border_style="thin",   color="BDBDBD")
    _thick = Side(border_style="medium", color=C_NAVY)

    def _border():
        return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _hborder():
        return Border(left=_thick, right=_thick, top=_thick, bottom=_thick)

    # ── Column layout ────────────────────────────────────────────────────────────────────────
    # A=spacer | B=Item ID | C=Description | D=Category | E=Qty |
    # F=Current Price | G=New Price [YELLOW] | H=Surcharge% [YELLOW] |
    # I=Change% [GREY formula] | J=Line Total [GREY formula] | K=Supplier
    COL_WIDTHS = {"A": 3, "B": 14, "C": 30, "D": 14, "E": 6,
                  "F": 16, "G": 16, "H": 13, "I": 13, "J": 16, "K": 14}

    HEADERS = [
        ("B", "Item ID"),
        ("C", "Description"),
        ("D", "Category"),
        ("E", "Qty"),
        ("F", "Current Price (€)"),
        ("G", "New Price (€)"),
        ("H", "Surcharge %"),
        ("I", "Change %"),
        ("J", "Line Total (€)"),
        ("K", "Supplier"),
    ]

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    SC_SHEET_NAMES = {"IN01": "\U0001f1ee\U0001f1f3 IN01", "NL07": "\U0001f1f3\U0001f1f1 NL07"}

    for sc in SUPPLY_CHAINS:
        sc_df = df[df["supply_chain"] == sc].copy()
        ws = wb.create_sheet(SC_SHEET_NAMES[sc])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5"

        for col, w in COL_WIDTHS.items():
            ws.column_dimensions[col].width = w

        # Row 1: spacer
        ws.row_dimensions[1].height = 6

        # Row 2: Banner
        ws.merge_cells("B2:K2")
        ws["B2"].value = f"QMS Component Price Database  —  {SC_LABELS.get(sc, sc)}  —  {today_str}"
        ws["B2"].font  = _hfont(size=13)
        ws["B2"].fill  = _fill(C_NAVY)
        ws["B2"].alignment = _center()
        ws.row_dimensions[2].height = 34

        # Row 3: Instruction
        ws.merge_cells("B3:K3")
        ws["B3"].value = ("Fill in column G (New Price) and H (Surcharge %). "
                          "Leave blank to keep current value. "
                          "Import via Cost Forge 2 → QMS Prices → Tab 4.")
        ws["B3"].font  = Font(italic=True, size=10, name="Calibri", color=C_NAVY)
        ws["B3"].fill  = _fill(C_LIGHT_BG)
        ws["B3"].alignment = _left()
        ws.row_dimensions[3].height = 20

        # Row 4: Column headers
        for col, label in HEADERS:
            is_yellow = col in ("G", "H")
            is_grey   = col in ("I", "J")
            ws[f"{col}4"].value     = label
            ws[f"{col}4"].font      = _hfont(size=9)
            ws[f"{col}4"].fill      = _fill(C_INPUT_BG if is_yellow else (C_LOCKED_BG if is_grey else C_NAVY))
            ws[f"{col}4"].alignment = _center()
            ws[f"{col}4"].border    = _border()
            if is_yellow or is_grey:
                ws[f"{col}4"].font = Font(bold=True, size=9, color=C_NAVY, name="Calibri")
        ws.row_dimensions[4].height = 24

        # Data rows — grouped by size
        sizes = sorted(sc_df["size_mm"].unique())
        current_row = 5
        first_data_row = None
        size_total_ranges: list[tuple[int, int, str]] = []  # (size, first_r, last_r)

        for size in sizes:
            size_df = (
                sc_df[sc_df["size_mm"] == size]
                .sort_values(["category", "item_id"])
                .reset_index(drop=True)
            )

            # Size subheader row
            ws.merge_cells(f"B{current_row}:K{current_row}")
            ws[f"B{current_row}"].value     = f"  MWJ-{size}"
            ws[f"B{current_row}"].font      = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
            ws[f"B{current_row}"].fill      = _fill(C_BLUE)
            ws[f"B{current_row}"].alignment = _left()
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            size_first = current_row

            for ridx, row in size_df.iterrows():
                r = current_row
                stripe = _fill(C_STRIPE) if ridx % 2 == 0 else _fill(C_WHITE)

                # Fixed / locked columns
                locked = [
                    ("B", str(row["item_id"]),    _left()),
                    ("C", str(row["description"]),_left()),
                    ("D", str(row["category"]),   _center()),
                    ("E", float(row["qty"]),       _right()),
                    ("F", float(row["price_eur"]), _right()),
                    ("K", str(row.get("supplier", "TBD")), _center()),
                ]
                for col, val, align in locked:
                    ws[f"{col}{r}"].value     = val
                    ws[f"{col}{r}"].font      = _font(size=10)
                    ws[f"{col}{r}"].fill      = stripe
                    ws[f"{col}{r}"].alignment = align
                    ws[f"{col}{r}"].border    = _border()
                    if col == "F":
                        ws[f"{col}{r}"].number_format = "€#,##0.00"
                    if col == "E":
                        ws[f"{col}{r}"].number_format = "0"

                # G: New Price (yellow / editable)
                ws[f"G{r}"].fill         = _fill(C_INPUT_BG)
                ws[f"G{r}"].border       = _border()
                ws[f"G{r}"].number_format = "€#,##0.00"
                ws[f"G{r}"].alignment    = _right()

                # H: Surcharge % (yellow / editable, pre-filled)
                ws[f"H{r}"].value        = float(row["surcharge_pct"])
                ws[f"H{r}"].fill         = _fill(C_INPUT_BG)
                ws[f"H{r}"].border       = _border()
                ws[f"H{r}"].number_format = "0.0"
                ws[f"H{r}"].alignment    = _right()

                # I: Change % formula (grey, locked)
                ws[f"I{r}"].value        = f'=IF(G{r}="","",IF(F{r}=0,"N/A",(G{r}-F{r})/F{r}))'
                ws[f"I{r}"].number_format = "+0.00%;-0.00%"
                ws[f"I{r}"].fill         = _fill(C_LOCKED_BG)
                ws[f"I{r}"].border       = _border()
                ws[f"I{r}"].alignment    = _right()

                # J: Line Total formula (grey, locked)
                ws[f"J{r}"].value        = (f'=IF(G{r}="",E{r}*F{r}*(1+H{r}/100),'
                                            f'E{r}*G{r}*(1+H{r}/100))')
                ws[f"J{r}"].number_format = "€#,##0.00"
                ws[f"J{r}"].fill         = _fill(C_LOCKED_BG)
                ws[f"J{r}"].border       = _border()
                ws[f"J{r}"].alignment    = _right()

                ws.row_dimensions[r].height = 20
                if first_data_row is None:
                    first_data_row = r
                current_row += 1

            size_last = current_row - 1
            size_total_ranges.append((size, size_first, size_last))

            # Total row for this size
            tr = current_row
            ws.merge_cells(f"B{tr}:I{tr}")
            ws[f"B{tr}"].value     = f"MWJ-{size} Total"
            ws[f"B{tr}"].font      = Font(bold=True, size=10, color=C_WHITE, name="Calibri")
            ws[f"B{tr}"].fill      = _fill(C_NAVY)
            ws[f"B{tr}"].alignment = _right()
            ws[f"B{tr}"].border    = _border()
            for col in list("CDEFGHI"):
                ws[f"{col}{tr}"].fill   = _fill(C_NAVY)
                ws[f"{col}{tr}"].border = _border()
            ws[f"J{tr}"].value        = f"=SUM(J{size_first}:J{size_last})"
            ws[f"J{tr}"].number_format = "€#,##0.00"
            ws[f"J{tr}"].font         = Font(bold=True, size=10, color=C_WHITE, name="Calibri")
            ws[f"J{tr}"].fill         = _fill(C_NAVY)
            ws[f"J{tr}"].alignment    = _right()
            ws[f"J{tr}"].border       = _border()
            ws[f"K{tr}"].fill   = _fill(C_NAVY)
            ws[f"K{tr}"].border = _border()
            ws.row_dimensions[tr].height = 22
            current_row += 1

            # Blank spacer
            ws.row_dimensions[current_row].height = 8
            current_row += 1

        if not sizes:
            ws[f"B5"].value = f"No data for {sc}."
            ws[f"B5"].font  = Font(italic=True, size=10, name="Calibri", color=C_NAVY)

    # Set first sheet active
    if wb.sheetnames:
        wb.active = wb[wb.sheetnames[0]]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_qms_excel(wb, df: pd.DataFrame) -> tuple[list[dict], pd.DataFrame]:
    """Read a filled QMS workbook and return (changes, updated_df)."""
    new_df = df.copy()

    # Build key map: (item_id, supply_chain, size_mm) -> df.index
    key_map: dict[tuple, int] = {}
    for idx, row in new_df.iterrows():
        key = (str(row["item_id"]).strip(), str(row["supply_chain"]).strip(), int(row["size_mm"]))
        key_map[key] = idx

    SC_SHEET_NAMES = {"IN01": "\U0001f1ee\U0001f1f3 IN01", "NL07": "\U0001f1f3\U0001f1f1 NL07"}
    changes: list[dict] = []

    def _safe_float(val):
        if val is None:
            return None
        try:
            return float(str(val).replace(",", "."))
        except Exception:
            return None

    for sc, sheet_name in SC_SHEET_NAMES.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        current_size: int | None = None

        for row in ws.iter_rows(min_row=5, values_only=True):
            b_val = row[1] if len(row) > 1 else None  # col B (0-indexed 1)
            if b_val is None:
                continue
            b_str = str(b_val).strip()

            # Detect size sub-header
            if b_str.startswith("MWJ-") or b_str.startswith("  MWJ-"):
                try:
                    current_size = int(b_str.strip().replace("MWJ-", ""))
                except ValueError:
                    pass
                continue

            # Skip total / blank rows
            if "Total" in b_str or b_str == "":
                continue

            if current_size is None:
                continue

            item_id = b_str
            key = (item_id, sc, current_size)
            if key not in key_map:
                continue

            idx = key_map[key]

            # col G = index 6 (B=1,C=2,D=3,E=4,F=5,G=6,H=7)
            g_val = row[6] if len(row) > 6 else None
            h_val = row[7] if len(row) > 7 else None

            new_price = _safe_float(g_val)
            new_surcharge = _safe_float(h_val)

            cur_price     = float(new_df.at[idx, "price_eur"])
            cur_surcharge = float(new_df.at[idx, "surcharge_pct"])

            if new_price is not None and abs(new_price - cur_price) > 1e-6:
                changes.append({
                    "key":   f"{sc}/{current_size}/{item_id}",
                    "field": "price_eur",
                    "old":   cur_price,
                    "new":   new_price,
                })
                new_df.at[idx, "price_eur"] = new_price
                new_df.at[idx, "valid_from"] = date.today().isoformat()

            if new_surcharge is not None and abs(new_surcharge - cur_surcharge) > 1e-6:
                changes.append({
                    "key":   f"{sc}/{current_size}/{item_id}",
                    "field": "surcharge_pct",
                    "old":   cur_surcharge,
                    "new":   new_surcharge,
                })
                new_df.at[idx, "surcharge_pct"] = new_surcharge
                new_df.at[idx, "valid_from"]     = date.today().isoformat()

    # Recompute line_total_eur
    new_df["line_total_eur"] = (
        new_df["qty"] * new_df["price_eur"] * (1 + new_df["surcharge_pct"] / 100)
    )
    return changes, new_df


# ─── Main UI ───────────────────────────────────────────────────────────────────────────────────

def main() -> None:
    home_button()
    st.title("\U0001f3d7️ QMS Component Price Database")
    st.caption(
        "Central price register for waterjet QMS components — "
        "supply chains IN01 (India) and NL07 (EU), all sizes."
    )

    df = load_qms()

    # ── KPI header cards ────────────────────────────────────────────────────────────────────
    combos = []
    if not df.empty:
        for sc in SUPPLY_CHAINS:
            for sz in sorted(df["size_mm"].unique()):
                sub = df[(df["supply_chain"] == sc) & (df["size_mm"] == sz)]
                if not sub.empty:
                    combos.append((sc, sz, sub["line_total_eur"].sum()))

    if combos:
        kpi_cols = st.columns(len(combos))
        for col, (sc, sz, total) in zip(kpi_cols, combos):
            col.metric(
                label=f"{SC_LABELS.get(sc, sc)} / MWJ-{sz}",
                value=f"€{total:,.0f}",
                help="Sum of qty × price × (1 + surcharge%) for this combination.",
            )
    else:
        st.warning(
            "No QMS price data loaded yet.\n\n"
            "**Getting started:** open the \'\U0001f4cb Price List\' tab "
            "and use the \'Add / Edit items\' expander to enter component prices, "
            "or import a filled Excel workbook via Tab 4."
        )

    st.divider()

    # ── 5 tabs ──────────────────────────────────────────────────────────────────────────────────
    tab_list, tab_adj, tab_gen, tab_imp, tab_hist = st.tabs([
        "\U0001f4cb Price List",
        "\U0001f4c8 Apply Surcharge / Adjustment",
        "\U0001f4e5 Generate Excel",
        "\U0001f4e4 Import Filled Excel",
        "\U0001f4dc History",
    ])

    # ════════════════════════════════════════════════════════════════════════════
    #  TAB 1: Price List
    # ════════════════════════════════════════════════════════════════════════════
    with tab_list:
        st.subheader("Component price list")

        if df.empty:
            st.info("No data — use the expander below to add items.")
        else:
            fcol1, fcol2, fcol3 = st.columns(3)
            with fcol1:
                sc_filter = st.selectbox(
                    "Supply chain", ["All"] + SUPPLY_CHAINS, key="pl_sc"
                )
            with fcol2:
                size_opts = ["All"] + [str(s) for s in sorted(df["size_mm"].unique())]
                sz_filter = st.selectbox("Size mm", size_opts, key="pl_sz")
            with fcol3:
                cat_filter = st.selectbox(
                    "Category", ["All"] + CATEGORIES, key="pl_cat"
                )

            fdf = df.copy()
            if sc_filter  != "All":
                fdf = fdf[fdf["supply_chain"] == sc_filter]
            if sz_filter  != "All":
                fdf = fdf[fdf["size_mm"] == int(sz_filter)]
            if cat_filter != "All":
                fdf = fdf[fdf["category"] == cat_filter]

            bom_total = fdf["line_total_eur"].sum()
            st.metric("Filtered BOM total", f"€{bom_total:,.2f}",
                      delta=f"{len(fdf)} line(s)")

            display_df = fdf[[
                "item_id", "description", "category", "supply_chain",
                "size_mm", "qty", "price_eur", "surcharge_pct", "line_total_eur",
            ]].copy()

            st.dataframe(
                display_df.style
                    .format({
                        "price_eur":     "€{:,.2f}",
                        "surcharge_pct": "{:.1f}%",
                        "line_total_eur":"€{:,.2f}",
                        "qty":           "{:.0f}",
                        "size_mm":       "{:d}",
                    }),
                use_container_width=True,
                hide_index=True,
            )

        # ── Add / Edit expander ──────────────────────────────────────────────────────────────
        with st.expander("➕ Add / Edit items", expanded=df.empty):
            st.caption(
                "Edit any cell below. "
                "Add new rows with the ➕ button at the bottom of the table. "
                "Click **Save changes** when done."
            )
            edit_df = df.drop(columns=["line_total_eur"], errors="ignore")
            edited = st.data_editor(
                edit_df,
                num_rows="dynamic",
                use_container_width=True,
                column_config={
                    "supply_chain": st.column_config.SelectboxColumn(
                        "Supply chain", options=SUPPLY_CHAINS, required=True
                    ),
                    "category": st.column_config.SelectboxColumn(
                        "Category", options=CATEGORIES
                    ),
                    "price_eur": st.column_config.NumberColumn(
                        "Price (€)", min_value=0.0, format="€%.2f"
                    ),
                    "surcharge_pct": st.column_config.NumberColumn(
                        "Surcharge %", min_value=0.0, format="%.1f%%"
                    ),
                    "qty": st.column_config.NumberColumn(
                        "Qty", min_value=0, step=1
                    ),
                    "size_mm": st.column_config.NumberColumn(
                        "Size mm", min_value=0, step=10
                    ),
                },
                key="qms_editor",
            )
            if st.button("\U0001f4be Save changes", type="primary", key="save_edit"):
                save_qms(edited)
                _log_update("edit", "Manual data-editor save", len(edited))
                st.success("Saved.")
                st.rerun()

    # ════════════════════════════════════════════════════════════════════════════
    #  TAB 2: Apply Surcharge / Adjustment
    # ════════════════════════════════════════════════════════════════════════════
    with tab_adj:
        st.subheader("Apply price % adjustment")
        st.markdown(
            "Select the scope below, enter a percentage change, "
            "and click **Apply**. "
            "Positive = price increase, negative = decrease. "
            "The `price_eur` column is updated; `valid_from` is set to today."
        )

        if df.empty:
            st.info("No data available. Add items in the Price List tab first.")
        else:
            ac1, ac2, ac3 = st.columns(3)
            with ac1:
                adj_sc = st.selectbox("Supply chain", ["All"] + SUPPLY_CHAINS, key="adj_sc")
            with ac2:
                adj_sz_opts = ["All"] + [str(s) for s in sorted(df["size_mm"].unique())]
                adj_sz = st.selectbox("Size mm", adj_sz_opts, key="adj_sz")
            with ac3:
                adj_cat = st.selectbox("Category", ["All"] + CATEGORIES, key="adj_cat")

            pct_change = st.number_input(
                "% change (e.g. 3.5 = +3.5%, -2 = -2%)",
                value=0.0, step=0.5, format="%.2f", key="adj_pct"
            )

            mask = pd.Series([True] * len(df), index=df.index)
            if adj_sc  != "All":
                mask &= df["supply_chain"] == adj_sc
            if adj_sz  != "All":
                mask &= df["size_mm"] == int(adj_sz)
            if adj_cat != "All":
                mask &= df["category"] == adj_cat

            n_items = mask.sum()
            st.info(f"**{n_items}** item(s) will be updated.")

            if st.button("▶️ Apply", type="primary", key="adj_apply", disabled=(n_items == 0 or pct_change == 0)):
                factor = 1 + pct_change / 100
                new_df = df.copy()
                new_df.loc[mask, "price_eur"] = (
                    new_df.loc[mask, "price_eur"] * factor
                ).round(2)
                new_df.loc[mask, "valid_from"] = date.today().isoformat()
                new_df["line_total_eur"] = (
                    new_df["qty"] * new_df["price_eur"] * (1 + new_df["surcharge_pct"] / 100)
                )
                save_qms(new_df)
                _log_update(
                    "price_adjustment",
                    f"sc={adj_sc} sz={adj_sz} cat={adj_cat} pct={pct_change:+.2f}%",
                    int(n_items),
                )
                st.success(f"✅ Applied {pct_change:+.2f}% to {n_items} item(s).")
                st.rerun()

    # ════════════════════════════════════════════════════════════════════════════
    #  TAB 3: Generate Excel
    # ════════════════════════════════════════════════════════════════════════════
    with tab_gen:
        st.subheader("Generate QMS price-update workbook")
        st.markdown(
            "Builds an Excel workbook with one sheet per supply chain. "
            "Yellow columns (G = New Price, H = Surcharge %) are editable; "
            "grey columns contain live formulas. "
            "Fill in the workbook then import it back via **Tab 4**."
        )

        if st.button("⚙️ Build QMS workbook", type="primary", key="gen_xlsx"):
            if df.empty:
                st.warning("No data to export — add items first.")
            else:
                with st.spinner("Building workbook …"):
                    try:
                        xlsx_bytes = _build_qms_excel(df)
                        st.session_state["qms_xlsx"] = xlsx_bytes
                        st.success(f"✅ Workbook ready — {len(xlsx_bytes)//1024} KB")
                    except Exception as exc:
                        import traceback
                        st.error(f"Build failed: {exc}")
                        st.code(traceback.format_exc(), language="python")
                        logger.exception("_build_qms_excel failed")

        if "qms_xlsx" in st.session_state:
            fname = f"qms_prices_{date.today().isoformat()}.xlsx"
            st.download_button(
                label=f"⬇️ Download {fname}",
                data=st.session_state["qms_xlsx"],
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # ════════════════════════════════════════════════════════════════════════════
    #  TAB 4: Import Filled Excel
    # ════════════════════════════════════════════════════════════════════════════
    with tab_imp:
        st.subheader("Import a filled QMS workbook")
        st.markdown(
            "Upload the workbook you filled in. "
            "Only non-blank values in column G (New Price) and H (Surcharge %) are applied. "
            "Use **Dry run** to preview changes before writing."
        )

        uploaded = st.file_uploader(
            "Upload qms_prices_*.xlsx",
            type=["xlsx"],
            help="Must be a workbook generated by the QMS Prices page.",
            key="qms_upload",
        )
        dry_run = st.toggle("\U0001f50d Dry run (preview only — no files written)", value=True, key="qms_dry")
        if dry_run:
            st.info("Dry run is ON — review the change list before applying.")

        if uploaded and st.button("▶️ Import", type="primary", key="qms_import"):
            try:
                from openpyxl import load_workbook as _lw
                wb = _lw(io.BytesIO(uploaded.read()), data_only=True)

                with st.spinner("Reading workbook …"):
                    changes, new_df = _read_qms_excel(wb, df)

                if not changes:
                    st.warning(
                        "No filled-in values detected. "
                        "Make sure you entered prices in column G of the workbook."
                    )
                else:
                    if dry_run:
                        st.success(
                            f"**Dry run** — {len(changes)} change(s) detected. "
                            "Disable dry run and re-run to apply."
                        )
                    else:
                        save_qms(new_df)
                        _log_update("excel_import", f"Uploaded: {uploaded.name}", len(changes))
                        st.success(f"✅ {len(changes)} change(s) applied.")
                        st.rerun()

                    st.markdown("**Change summary:**")
                    for ch in changes:
                        old_v = ch["old"]
                        new_v = ch["new"]
                        arrow = "\U0001f53a" if new_v > old_v else "\U0001f53b"
                        st.markdown(
                            f"- `{ch['key']}` / {ch['field']}: "
                            f"**{old_v:.4f}** → **{new_v:.4f}** {arrow}"
                        )

            except Exception as exc:
                st.error(f"Import failed: {exc}")
                logger.exception("_read_qms_excel failed")

    # ════════════════════════════════════════════════════════════════════════════
    #  TAB 5: History
    # ════════════════════════════════════════════════════════════════════════════
    with tab_hist:
        st.subheader("Update history")
        if not QMS_LOG.exists():
            st.info(
                "No history yet. "
                "Log entries are created whenever prices are adjusted or imported."
            )
        else:
            try:
                log_df = pd.read_csv(QMS_LOG)
                if log_df.empty:
                    st.info("Log file exists but has no entries yet.")
                else:
                    st.dataframe(
                        log_df.sort_values("date", ascending=False).reset_index(drop=True),
                        use_container_width=True,
                        hide_index=True,
                    )
            except Exception as exc:
                st.error(f"Could not read log: {exc}")


guard(main)
