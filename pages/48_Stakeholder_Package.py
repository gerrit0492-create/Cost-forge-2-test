"""
48_Stakeholder_Package.py  —  Stakeholder Package Generator
Generates a professional multi-tab Excel workbook + multi-section PDF
covering the full project workflow for stakeholder communication.
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime

import pandas as pd
import streamlit as st

from utils.completeness import WATERJET_SUBSYSTEMS
from utils.currency import fmt
from utils.io import (
    load_bom,
    load_change_orders,
    load_cost_timeline,
    load_escalation,
    load_india_lc,
    load_materials,
    load_milestones,
    load_nre,
    load_outbound,
    load_processes,
    load_quotes,
    load_risk,
    load_transport,
)
from utils.nav import home_button
from utils.pricing import compute_costs
from utils.project import load_project_meta
from utils.quotes import apply_best_quotes
from utils.safe import guard
from utils.style import inject_css, page_header

# ── openpyxl ────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── reportlab ────────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ── Constants ────────────────────────────────────────────────────────────────
TODAY = date.today()
TODAY_STR = TODAY.isoformat()
EMISSION_FACTOR = 0.233  # kg CO2e / kWh (EU average grid)

NAVY = "#1F3864"
NAVY_FILL = "1F3864"
LIGHT_BLUE_FILL = "DCE6F1"
ORANGE_FILL = "FFC000"
WHITE = "FFFFFF"

_xl = lambda s: re.sub(r"[^\x00-\x7F]|[\U0001F300-\U0001FFFF]", "", str(s)).strip()  # noqa: E731


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING (cached)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=30)
def _load_all():
    bom = load_bom()
    mats = load_materials()
    procs = load_processes()
    quotes = load_quotes()
    transport = load_transport()
    outbound = load_outbound()
    nre = load_nre()
    risk = load_risk()
    escalation = load_escalation()
    milestones = load_milestones()
    cost_timeline = load_cost_timeline()
    change_orders = load_change_orders()
    india_lc = load_india_lc()
    meta = load_project_meta()

    mats_quoted = apply_best_quotes(mats, quotes)
    # Use default energy rate (0.20 EUR/kWh) for correct EUR financials
    df = compute_costs(mats_quoted, procs, bom)
    # Add kWh column: divide energy_cost back out by the default rate
    df["energy_kwh"] = df["energy_cost"] / 0.20

    return {
        "bom": bom,
        "mats": mats,
        "procs": procs,
        "quotes": quotes,
        "transport": transport,
        "outbound": outbound,
        "nre": nre,
        "risk": risk,
        "escalation": escalation,
        "milestones": milestones,
        "cost_timeline": cost_timeline,
        "change_orders": change_orders,
        "india_lc": india_lc,
        "meta": meta,
        "df": df,
    }


# ══════════════════════════════════════════════════════════════════════════════
# SUBSYSTEM HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _assign_subsystem(line_id: str) -> str:
    """Return subsystem prefix for a line_id (longest-first match)."""
    upper = str(line_id).upper()
    for prefix in sorted(WATERJET_SUBSYSTEMS.keys(), key=len, reverse=True):
        if upper.startswith(prefix):
            return prefix
    return "OTHER"


def _build_subsystem_table(df: pd.DataFrame, bom: pd.DataFrame) -> pd.DataFrame:
    """Build per-subsystem aggregated cost table."""
    merged = df.copy()
    if "line_id" not in merged.columns and "line_id" in bom.columns:
        merged["line_id"] = bom["line_id"].values

    merged["_scope"] = merged["line_id"].apply(_assign_subsystem)

    rows = []
    all_prefixes = sorted(WATERJET_SUBSYSTEMS.keys(), key=len, reverse=True)
    for prefix in all_prefixes:
        info = WATERJET_SUBSYSTEMS[prefix]
        sub = merged[merged["_scope"] == prefix]
        if sub.empty:
            continue
        mat = sub["material_cost"].sum()
        proc = sub["process_cost"].sum()
        oh = sub["overhead"].sum()
        marg = sub["margin"].sum()
        sell = sub["total_cost"].sum()
        mass = sub["mass_kg"].sum() if "mass_kg" in sub.columns else 0.0
        kwh = sub["energy_kwh"].sum() if "energy_kwh" in sub.columns else sub["energy_cost"].sum() / 0.20
        rows.append({
            "prefix": prefix,
            "scope": info["name"],
            "lines": len(sub),
            "mass_kg": mass,
            "material_eur": mat,
            "process_eur": proc,
            "overhead_eur": oh,
            "margin_eur": marg,
            "sell_eur": sell,
            "kwh": kwh,
            "co2e_kg": kwh * EMISSION_FACTOR,
        })

    result = pd.DataFrame(rows)
    if not result.empty:
        total_sell = result["sell_eur"].sum()
        result["share_pct"] = result["sell_eur"] / total_sell * 100 if total_sell else 0.0
        result["eur_per_kg"] = result.apply(
            lambda r: r["sell_eur"] / r["mass_kg"] if r["mass_kg"] > 0 else 0.0, axis=1
        )
    return result


def _build_risks(data: dict) -> list[dict]:
    """Auto-generate risks from data quality + loaded risk register."""
    risks = []
    df = data["df"]
    quotes = data["quotes"]
    meta = data["meta"]

    # Check expired quotes
    if not quotes.empty and "valid_until" in quotes.columns:
        today_ts = pd.Timestamp.today().normalize()
        expired = quotes[pd.to_datetime(quotes["valid_until"], errors="coerce") < today_ts]
        n_exp = len(expired)
        if n_exp > 0:
            risks.append({
                "Category": "Procurement",
                "Title": f"{n_exp} supplier quote(s) expired",
                "Severity": "High" if n_exp >= 3 else "Medium",
                "Impact EUR": df["material_cost"].sum() * 0.05,
                "Mitigation": "Refresh supplier quotes before finalising proposal",
                "Owner": "",
                "Status": "Open",
            })

    # Low margin
    if not df.empty:
        base = df["base_cost"].sum()
        marg = df["margin"].sum()
        sell = df["total_cost"].sum()
        margin_pct = marg / sell if sell else 0.0
        if margin_pct < 0.10:
            risks.append({
                "Category": "Commercial",
                "Title": f"Low margin: {margin_pct*100:.1f}%",
                "Severity": "High",
                "Impact EUR": (0.10 - margin_pct) * sell,
                "Mitigation": "Review process rates or negotiate material pricing",
                "Owner": "",
                "Status": "Open",
            })

        # Material share
        mat_share = df["material_cost"].sum() / base if base else 0.0
        if mat_share > 0.75:
            risks.append({
                "Category": "Cost Structure",
                "Title": f"High material share: {mat_share*100:.0f}%",
                "Severity": "Medium",
                "Impact EUR": df["material_cost"].sum() * 0.05,
                "Mitigation": "Hedge commodity exposure; fix prices with long-term agreements",
                "Owner": "",
                "Status": "Open",
            })

    # Append risk register rows
    risk_df = data.get("risk", pd.DataFrame())
    if isinstance(risk_df, pd.DataFrame) and not risk_df.empty:
        for _, row in risk_df.iterrows():
            risks.append({
                "Category": str(row.get("category", "General")),
                "Title": str(row.get("title", row.get("risk", ""))),
                "Severity": str(row.get("severity", "Medium")),
                "Impact EUR": float(row.get("impact_eur", 0) or 0),
                "Mitigation": str(row.get("mitigation", "")),
                "Owner": str(row.get("owner", "")),
                "Status": str(row.get("status", "Open")),
            })

    return risks


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WORKBOOK BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def _xl_header_fill():
    return PatternFill("solid", fgColor=NAVY_FILL)

def _xl_alt_fill():
    return PatternFill("solid", fgColor=LIGHT_BLUE_FILL)

def _xl_orange_fill():
    return PatternFill("solid", fgColor=ORANGE_FILL)

def _xl_header_font():
    return Font(bold=True, color=WHITE, size=10)

def _xl_bold():
    return Font(bold=True, size=10)

def _xl_normal():
    return Font(size=10)

def _xl_center():
    return Alignment(horizontal="center", vertical="center")

def _xl_left():
    return Alignment(horizontal="left", vertical="center")

def _xl_right():
    return Alignment(horizontal="right", vertical="center")

def _apply_header_row(ws, row_num: int, values: list, col_widths: dict | None = None):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=ci, value=_xl(val))
        cell.fill = _xl_header_fill()
        cell.font = _xl_header_font()
        cell.alignment = _xl_center()
    ws.freeze_panes = f"A{row_num + 1}"

def _apply_data_row(ws, row_num: int, values: list, alternate: bool = False, fmt_map: dict | None = None):
    fill = _xl_alt_fill() if alternate else None
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=ci, value=val if not isinstance(val, str) else _xl(val))
        if fill:
            cell.fill = fill
        cell.font = _xl_normal()
        cell.alignment = _xl_left()
        if fmt_map and ci in fmt_map:
            cell.number_format = fmt_map[ci]

def _auto_col_widths(ws, widths: list[int]):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _sheet_cover(wb: Workbook, data: dict, company: str, prepared_by: str, confidentiality: str):
    ws = wb.create_sheet("COVER")
    ws.sheet_properties.tabColor = "1F3864"

    df = data["df"]
    meta = data["meta"]
    bom = data["bom"]
    mats = data["mats"]

    proj_name = _xl(meta.get("name", "Unnamed Project"))
    customer = _xl(meta.get("customer", ""))
    ref = _xl(meta.get("estimate_ref", meta.get("ref", "")))
    maturity = _xl(meta.get("maturity", "Budget (±15%)"))

    sell = df["total_cost"].sum()
    base = df["base_cost"].sum()
    margin = df["margin"].sum()
    margin_pct = margin / sell * 100 if sell else 0.0
    mat_cost = df["material_cost"].sum()
    mat_share = mat_cost / base * 100 if base else 0.0
    total_mass = df["mass_kg"].sum() if "mass_kg" in df.columns else 0.0
    eur_per_kg = sell / total_mass if total_mass > 0 else 0.0
    bom_lines = len(bom)
    unique_mats = bom["material_id"].nunique() if "material_id" in bom.columns else 0

    _apply_header_row(ws, 1, ["Field", "Value"])

    rows_data = [
        ("Project name", proj_name),
        ("Customer", customer),
        ("Estimate reference", ref),
        ("Date generated", TODAY_STR),
        ("Estimate maturity", maturity),
        ("Report prepared by", _xl(prepared_by)),
        ("Confidentiality", _xl(confidentiality)),
        ("Company", _xl(company)),
        ("---", "---"),
        ("Sell price (EUR)", sell),
        ("Base cost (EUR)", base),
        ("Margin (EUR)", margin),
        ("Margin (%)", margin_pct / 100),
        ("Material share (%)", mat_share / 100),
        ("EUR per kg", eur_per_kg),
        ("Total mass (kg)", total_mass),
        ("BOM lines", bom_lines),
        ("Unique materials", unique_mats),
        ("---", "---"),
        ("SCOPE", "This report covers the full cost engineering workflow for the above project, "
                  "including BOM costs, procurement status, energy and carbon, "
                  "subsystem breakdowns, India local content, and risk register."),
    ]

    curr_fmt = '#,##0.00'
    pct_fmt = '0.0%'
    for ri, (field, val) in enumerate(rows_data, 2):
        alt = (ri % 2 == 0)
        cell_f = ws.cell(row=ri, column=1, value=_xl(field))
        cell_v = ws.cell(row=ri, column=2, value=val if not isinstance(val, str) else _xl(val))
        fill = _xl_alt_fill() if alt else None
        if fill:
            cell_f.fill = fill
            cell_v.fill = fill
        cell_f.font = _xl_bold() if field.startswith("---") else _xl_normal()
        cell_v.font = _xl_normal()
        cell_f.alignment = _xl_left()
        cell_v.alignment = _xl_left()
        if field in ("Sell price (EUR)", "Base cost (EUR)", "Margin (EUR)", "EUR per kg"):
            cell_v.number_format = curr_fmt
        elif field in ("Margin (%)", "Material share (%)"):
            cell_v.number_format = pct_fmt

    _auto_col_widths(ws, [28, 60])


def _sheet_exec_summary(wb: Workbook, data: dict):
    ws = wb.create_sheet("EXECUTIVE SUMMARY")
    ws.sheet_properties.tabColor = "2E75B6"

    df = data["df"]
    bom = data["bom"]

    sell = df["total_cost"].sum()
    base = df["base_cost"].sum()
    mat = df["material_cost"].sum()
    proc = df["process_cost"].sum()
    mach = df["machine_cost"].sum() if "machine_cost" in df.columns else 0.0
    lab = df["labour_cost"].sum() if "labour_cost" in df.columns else 0.0
    tool = df["tooling_cost"].sum() if "tooling_cost" in df.columns else 0.0
    energy_eur = df["energy_cost"].sum()
    rework = df["rework_cost"].sum() if "rework_cost" in df.columns else 0.0
    moq = df["moq_excess_cost"].sum() if "moq_excess_cost" in df.columns else 0.0
    oh = df["overhead"].sum()
    marg = df["margin"].sum()
    kwh_total = df["energy_kwh"].sum() if "energy_kwh" in df.columns else energy_eur / 0.20

    ws.cell(row=1, column=1, value="EXECUTIVE SUMMARY — COST KPIs").font = Font(bold=True, size=13, color=NAVY_FILL)
    _apply_header_row(ws, 2, ["Cost Element", "EUR", "% of Base Cost", "Note"])

    pattern = df["pattern_cost"].sum() if "pattern_cost" in df.columns else 0.0

    kpi_rows = [
        ("Material cost",           mat,        mat / base * 100 if base else 0, ""),
        ("MOQ excess cost",         moq,        moq / base * 100 if base else 0, "Min order qty premium — often exceeds material cost"),
        ("Pattern / tooling NRE",   pattern,    pattern / base * 100 if base else 0, "Casting pattern amortised per unit"),
        ("Process cost",            proc,       proc / base * 100 if base else 0, ""),
        ("  — Machine cost",        mach,       mach / base * 100 if base else 0, "incl. in Process"),
        ("  — Labour cost",         lab,        lab / base * 100 if base else 0, "incl. in Process"),
        ("  — Tooling cost",        tool,       tool / base * 100 if base else 0, "incl. in Process"),
        ("  — Energy cost",         energy_eur, energy_eur / base * 100 if base else 0, f"{kwh_total:,.0f} kWh @ 0.20 EUR/kWh"),
        ("  — Rework cost",         rework,     rework / base * 100 if base else 0, "incl. in Process"),
        ("Overhead",                oh,         oh / base * 100 if base else 0, ""),
        ("Base cost",               base,       100.0, ""),
        ("Margin",                  marg,       marg / base * 100 if base else 0, ""),
        ("Sell price",              sell,       sell / base * 100 if base else 0, ""),
    ]
    for ri, (label, eur, pct, note) in enumerate(kpi_rows, 3):
        alt = (ri % 2 == 0)
        bold_row = label in ("Base cost", "Sell price")
        sub_row = label.startswith("  —")
        fill = _xl_alt_fill() if alt else None
        for ci, val in enumerate([_xl(label), eur, pct / 100, _xl(note)], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.font = _xl_bold() if bold_row else (Font(italic=True, size=9) if sub_row else _xl_normal())
            cell.alignment = _xl_left()
            if ci == 2:
                cell.number_format = '#,##0.00'
            elif ci == 3:
                cell.number_format = '0.0%'
            elif ci == 4:
                cell.font = Font(italic=True, size=9, color="666666")

    # Subsystem breakdown
    sub_df = _build_subsystem_table(df, bom)
    offset = len(kpi_rows) + 5
    ws.cell(row=offset, column=1, value="SUBSYSTEM BREAKDOWN").font = Font(bold=True, size=11, color=NAVY_FILL)
    _apply_header_row(ws, offset + 1, ["Scope", "Sell price EUR", "Share %", "Mass kg", "EUR/kg", "kWh"])

    if not sub_df.empty:
        for ri, row in enumerate(sub_df.itertuples(), offset + 2):
            alt = (ri % 2 == 0)
            fill = _xl_alt_fill() if alt else None
            vals = [_xl(row.scope), row.sell_eur, row.share_pct / 100, row.mass_kg, row.eur_per_kg, row.kwh]
            fmt_map = {2: '#,##0.00', 3: '0.0%', 4: '#,##0.00', 5: '#,##0.00', 6: '#,##0.0'}
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
                cell.font = _xl_normal()
                cell.alignment = _xl_left()
                if ci in fmt_map:
                    cell.number_format = fmt_map[ci]

    # Col widths: A=Cost Element/Scope, B=EUR, C=%, D=Note, E=EUR/kg, F=kWh
    _auto_col_widths(ws, [34, 16, 14, 38, 12, 12])


def _sheet_waterfall(wb: Workbook, data: dict):
    ws = wb.create_sheet("COST WATERFALL")
    ws.sheet_properties.tabColor = "4472C4"

    df = data["df"]
    sell = df["total_cost"].sum()

    def _pct(v):
        return v / sell if sell else 0.0

    moq = df["moq_excess_cost"].sum() if "moq_excess_cost" in df.columns else 0.0
    pattern_wf = df["pattern_cost"].sum() if "pattern_cost" in df.columns else 0.0
    energy_eur = df["energy_cost"].sum()
    kwh_total = df["energy_kwh"].sum() if "energy_kwh" in df.columns else energy_eur / 0.20

    elements = [
        ("Material cost", df["material_cost"].sum()),
        ("MOQ excess cost", moq),
        ("Pattern / tooling NRE", pattern_wf),
        ("Process cost", df["process_cost"].sum()),
        ("  Machine cost", df["machine_cost"].sum() if "machine_cost" in df.columns else 0.0),
        ("  Labour cost", df["labour_cost"].sum() if "labour_cost" in df.columns else 0.0),
        ("  Tooling cost", df["tooling_cost"].sum() if "tooling_cost" in df.columns else 0.0),
        (f"  Energy cost ({kwh_total:,.0f} kWh)", energy_eur),
        ("  Rework cost", df["rework_cost"].sum() if "rework_cost" in df.columns else 0.0),
        ("Overhead", df["overhead"].sum()),
        ("Base cost", df["base_cost"].sum()),
        ("Margin", df["margin"].sum()),
        ("Sell price", sell),
    ]

    _apply_header_row(ws, 1, ["Cost Element", "EUR", "% of Sell Price"])
    for ri, (label, eur) in enumerate(elements, 2):
        alt = (ri % 2 == 0)
        bold_row = label in ("Base cost", "Sell price")
        fill = _xl_alt_fill() if alt else None
        for ci, val in enumerate([_xl(label), eur, _pct(eur)], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.font = _xl_bold() if bold_row else _xl_normal()
            cell.alignment = _xl_left()
            if ci == 2:
                cell.number_format = '#,##0.00'
            elif ci == 3:
                cell.number_format = '0.0%'

    _auto_col_widths(ws, [24, 16, 16])


def _sheet_subsystem(wb: Workbook, data: dict):
    ws = wb.create_sheet("SUBSYSTEM BREAKDOWN")
    ws.sheet_properties.tabColor = "70AD47"

    df = data["df"]
    bom = data["bom"]
    sub_df = _build_subsystem_table(df, bom)

    headers = ["Scope", "Lines", "Mass (kg)", "Material (EUR)", "Process (EUR)",
               "Overhead (EUR)", "Margin (EUR)", "Sell price (EUR)", "Share %", "EUR/kg", "kWh", "CO2e kg"]
    _apply_header_row(ws, 1, headers)

    if not sub_df.empty:
        for ri, row in enumerate(sub_df.itertuples(), 2):
            alt = (ri % 2 == 0)
            fill = _xl_alt_fill() if alt else None
            co2e = row.kwh * EMISSION_FACTOR
            vals = [_xl(row.scope), row.lines, row.mass_kg, row.material_eur, row.process_eur,
                    row.overhead_eur, row.margin_eur, row.sell_eur, row.share_pct / 100,
                    row.eur_per_kg, row.kwh, co2e]
            fmt_map = {3: '#,##0.00', 4: '#,##0.00', 5: '#,##0.00', 6: '#,##0.00',
                       7: '#,##0.00', 8: '#,##0.00', 9: '0.0%', 10: '#,##0.00',
                       11: '#,##0.0', 12: '#,##0.0'}
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
                cell.font = _xl_normal()
                cell.alignment = _xl_left()
                if ci in fmt_map:
                    cell.number_format = fmt_map[ci]

    _auto_col_widths(ws, [26, 8, 12, 15, 14, 14, 13, 16, 10, 10, 10, 10])


def _sheet_bom_detail(wb: Workbook, data: dict):
    ws = wb.create_sheet("BOM DETAIL")
    ws.sheet_properties.tabColor = "ED7D31"

    df = data["df"]
    bom = data["bom"]

    headers = ["Line ID", "Component", "Material", "Qty", "Mass kg",
               "Material cost", "Process cost", "Overhead", "Margin", "Total cost", "kWh"]
    _apply_header_row(ws, 1, headers)

    curr_fmt = '#,##0.00'
    kwh_fmt = '#,##0.0'

    for ri, row in enumerate(df.itertuples(), 2):
        alt = (ri % 2 == 0)
        fill = _xl_alt_fill() if alt else None
        line_id = _xl(getattr(row, "line_id", ""))
        part = _xl(getattr(row, "part_name", bom.iloc[ri - 2]["part_name"] if ri - 2 < len(bom) else ""))
        mat_id = _xl(getattr(row, "material_id", ""))
        qty = getattr(row, "qty", 0)
        mass = getattr(row, "mass_kg", 0.0)
        mat_c = getattr(row, "material_cost", 0.0)
        proc_c = getattr(row, "process_cost", 0.0)
        oh = getattr(row, "overhead", 0.0)
        marg = getattr(row, "margin", 0.0)
        tot = getattr(row, "total_cost", 0.0)
        kwh = getattr(row, "energy_cost", 0.0)

        vals = [line_id, part, mat_id, qty, mass, mat_c, proc_c, oh, marg, tot, kwh]
        fmt_map = {5: curr_fmt, 6: curr_fmt, 7: curr_fmt, 8: curr_fmt,
                   9: curr_fmt, 10: curr_fmt, 11: kwh_fmt}
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.font = _xl_normal()
            cell.alignment = _xl_left()
            if ci in fmt_map:
                cell.number_format = fmt_map[ci]

    _auto_col_widths(ws, [14, 28, 14, 6, 10, 14, 14, 12, 12, 14, 10])


def _sheet_procurement(wb: Workbook, data: dict):
    ws = wb.create_sheet("PROCUREMENT")
    ws.sheet_properties.tabColor = "FF0000"

    quotes = data["quotes"]
    today_ts = pd.Timestamp.today().normalize()

    headers = ["Supplier", "Material ID", "Price EUR/kg", "Lead time days", "Valid until", "Preferred", "Status"]
    _apply_header_row(ws, 1, headers)

    if not quotes.empty:
        for ri, row in enumerate(quotes.itertuples(), 2):
            alt = (ri % 2 == 0)
            try:
                valid_dt = pd.to_datetime(row.valid_until, errors="coerce")
                expired = bool(valid_dt < today_ts) if not pd.isnull(valid_dt) else False
            except Exception:
                expired = False
            status = "EXPIRED" if expired else "ACTIVE"
            fill = _xl_orange_fill() if expired else (_xl_alt_fill() if alt else None)
            valid_str = valid_dt.strftime("%Y-%m-%d") if not pd.isnull(valid_dt) else ""
            vals = [_xl(row.supplier), _xl(row.material_id), row.price_eur_per_kg,
                    int(row.lead_time_days) if not pd.isnull(row.lead_time_days) else "",
                    valid_str, "Yes" if row.preferred else "No", status]
            fmt_map = {3: '#,##0.00'}
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
                cell.font = _xl_normal()
                cell.alignment = _xl_left()
                if ci in fmt_map:
                    cell.number_format = fmt_map[ci]

    _auto_col_widths(ws, [22, 16, 14, 16, 14, 10, 10])


def _sheet_energy_carbon(wb: Workbook, data: dict):
    ws = wb.create_sheet("ENERGY & CARBON")
    ws.sheet_properties.tabColor = "00B050"

    df = data["df"]
    bom = data["bom"]
    sub_df = _build_subsystem_table(df, bom)

    ws.cell(row=1, column=1, value="SECTION A — Energy & Carbon by Subsystem").font = Font(bold=True, size=11, color=NAVY_FILL)
    headers = ["Scope", "kWh", "CO2e kg", "CO2e t", "kWh/EUR intensity"]
    _apply_header_row(ws, 2, headers)

    total_kwh = 0.0
    total_co2 = 0.0

    if not sub_df.empty:
        for ri, row in enumerate(sub_df.itertuples(), 3):
            alt = (ri % 2 == 0)
            fill = _xl_alt_fill() if alt else None
            co2_kg = row.kwh * EMISSION_FACTOR
            co2_t = co2_kg / 1000
            intensity = row.kwh / row.sell_eur if row.sell_eur > 0 else 0.0
            total_kwh += row.kwh
            total_co2 += co2_kg
            vals = [_xl(row.scope), row.kwh, co2_kg, co2_t, intensity]
            fmt_map = {2: '#,##0.0', 3: '#,##0.0', 4: '#,##0.000', 5: '#,##0.000'}
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
                cell.font = _xl_normal()
                cell.alignment = _xl_left()
                if ci in fmt_map:
                    cell.number_format = fmt_map[ci]

    offset = (len(sub_df) if not sub_df.empty else 0) + 5
    ws.cell(row=offset, column=1, value="SECTION B — Summary").font = Font(bold=True, size=11, color=NAVY_FILL)
    summary_rows = [
        ("Total kWh", total_kwh),
        ("Total CO2e kg", total_kwh * EMISSION_FACTOR),
        ("Total CO2e t", total_kwh * EMISSION_FACTOR / 1000),
        ("Emission factor", EMISSION_FACTOR),
    ]
    for ri2, (label, val) in enumerate(summary_rows, offset + 1):
        ws.cell(row=ri2, column=1, value=_xl(label)).font = _xl_bold()
        cell_v = ws.cell(row=ri2, column=2, value=val)
        cell_v.font = _xl_normal()
        cell_v.number_format = '#,##0.000'

    note_row = offset + len(summary_rows) + 2
    ws.cell(row=note_row, column=1,
            value="Note: Scope 2 emissions (purchased electricity for manufacturing). EU average grid 0.233 kg CO2e/kWh").font = Font(italic=True, size=9, color="666666")

    _auto_col_widths(ws, [30, 14, 14, 12, 18])


def _sheet_india_lc(wb: Workbook, data: dict):
    ws = wb.create_sheet("INDIA LOCAL CONTENT")
    ws.sheet_properties.tabColor = "FF8C00"

    india_lc = data["india_lc"]
    meta = data["meta"]

    headers = ["Scope", "Origin", "IC Fraction", "Indian Manufacturer", "HS Code", "Declaration Rxd?", "Ref/Date"]
    _apply_header_row(ws, 1, headers)

    scope_rows = india_lc[india_lc["line_id"].str.startswith("SCOPE_")] if not india_lc.empty and "line_id" in india_lc.columns else pd.DataFrame()

    if scope_rows.empty:
        ws.cell(row=2, column=1, value="No India Local Content data entered yet. Use India Local Content page.")
        ws.merge_cells("A2:G2")
        ws.cell(row=2, column=1).font = Font(italic=True, color="888888")
    else:
        for ri, row in enumerate(scope_rows.itertuples(), 2):
            alt = (ri % 2 == 0)
            fill = _xl_alt_fill() if alt else None
            vals = [_xl(row.line_id), _xl(row.origin), float(row.ic_value_pct or 0),
                    _xl(row.indian_supplier), _xl(row.hs_code),
                    _xl(row.declaration_rxd), _xl(row.declaration_ref)]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
                cell.font = _xl_normal()
                cell.alignment = _xl_left()
                if ci == 3:
                    cell.number_format = '0.0%'

        # Summary row
        sum_row = len(scope_rows) + 3
        avg_ic = scope_rows["ic_value_pct"].mean() if "ic_value_pct" in scope_rows.columns else 0.0
        req_ic = meta.get("required_ic_pct", "")
        ws.cell(row=sum_row, column=1, value="Summary").font = _xl_bold()
        ws.cell(row=sum_row, column=2, value=f"Declared IC%: {avg_ic*100:.1f}%")
        ws.cell(row=sum_row, column=3, value=f"Required IC%: {req_ic if req_ic else '(not set)'}")

    _auto_col_widths(ws, [20, 16, 12, 26, 14, 16, 20])


def _sheet_risk(wb: Workbook, data: dict):
    ws = wb.create_sheet("RISK & ACTIONS")
    ws.sheet_properties.tabColor = "C00000"

    risks = _build_risks(data)
    headers = ["Category", "Title", "Severity", "Impact EUR", "Mitigation", "Owner", "Status"]
    _apply_header_row(ws, 1, headers)

    sev_colors = {"High": "C00000", "Medium": "FFC000", "Low": "70AD47"}

    for ri, risk in enumerate(risks, 2):
        alt = (ri % 2 == 0)
        fill = _xl_alt_fill() if alt else None
        sev = risk.get("Severity", "Medium")
        vals = [_xl(risk["Category"]), _xl(risk["Title"]), _xl(sev),
                risk.get("Impact EUR", 0), _xl(risk.get("Mitigation", "")),
                _xl(risk.get("Owner", "")), _xl(risk.get("Status", "Open"))]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 3:  # severity
                cell.fill = PatternFill("solid", fgColor=sev_colors.get(sev, LIGHT_BLUE_FILL))
                cell.font = Font(bold=True, color=WHITE if sev == "High" else "000000", size=10)
            else:
                if fill:
                    cell.fill = fill
                cell.font = _xl_normal()
            cell.alignment = _xl_left()
            if ci == 4:
                cell.number_format = '#,##0.00'

    _auto_col_widths(ws, [16, 34, 10, 14, 40, 14, 10])


def build_excel(data: dict, company: str, prepared_by: str, confidentiality: str) -> bytes:
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_cover(wb, data, company, prepared_by, confidentiality)
    _sheet_exec_summary(wb, data)
    _sheet_waterfall(wb, data)
    _sheet_subsystem(wb, data)
    _sheet_bom_detail(wb, data)
    _sheet_procurement(wb, data)
    _sheet_energy_carbon(wb, data)
    _sheet_india_lc(wb, data)
    _sheet_risk(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PDF BUILDER
# ══════════════════════════════════════════════════════════════════════════════

PDF_NAVY = colors.HexColor("#1F3864")
PDF_DARK = colors.HexColor("#404040")
PDF_LIGHT = colors.HexColor("#DCE6F1")
PDF_ORANGE = colors.HexColor("#FFC000")
PDF_WHITE = colors.white
PDF_RED = colors.HexColor("#C00000")
PDF_GREEN = colors.HexColor("#70AD47")


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        "CoverTitle",
        fontSize=28, fontName="Helvetica-Bold",
        textColor=PDF_NAVY, alignment=1, spaceAfter=12,
    ))
    styles.add(ParagraphStyle(
        "CoverSub",
        fontSize=13, fontName="Helvetica",
        textColor=PDF_DARK, alignment=1, spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        "SectionHeading",
        fontSize=14, fontName="Helvetica-Bold",
        textColor=PDF_NAVY, spaceBefore=14, spaceAfter=8,
        borderPad=4,
    ))
    styles.add(ParagraphStyle(
        "BodySmall",
        fontSize=9, fontName="Helvetica",
        textColor=PDF_DARK, spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        "NoteItalic",
        fontSize=8, fontName="Helvetica-Oblique",
        textColor=colors.HexColor("#888888"), spaceAfter=4,
    ))
    return styles


def _pdf_table_style(has_total_row: bool = False, total_row_idx: int = -1) -> TableStyle:
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [PDF_LIGHT, PDF_WHITE]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFC9D0")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    if has_total_row and total_row_idx > 0:
        cmds += [
            ("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
            ("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), PDF_LIGHT),
            ("LINEABOVE", (0, total_row_idx), (-1, total_row_idx), 1, PDF_NAVY),
        ]
    return TableStyle(cmds)


def _page_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#888888"))
    canvas.drawString(2 * cm, 1.2 * cm, f"Cost Forge 2 — Stakeholder Report — Generated {TODAY_STR}")
    canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf(data: dict, company: str, prepared_by: str, confidentiality: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2.5 * cm,
    )
    styles = _pdf_styles()
    story = []

    df = data["df"]
    bom = data["bom"]
    meta = data["meta"]
    quotes = data["quotes"]

    proj_name = meta.get("name", "Unnamed Project")
    customer = meta.get("customer", "")
    maturity = meta.get("maturity", "Budget (±15%)")

    sell = df["total_cost"].sum()
    base = df["base_cost"].sum()
    margin = df["margin"].sum()
    total_mass = df["mass_kg"].sum() if "mass_kg" in df.columns else 0.0

    # ── Section 1: Cover ──────────────────────────────────────────────────────
    story.append(Spacer(1, 2 * cm))
    story.append(Paragraph(proj_name, styles["CoverTitle"]))
    story.append(HRFlowable(width="100%", thickness=2, color=PDF_NAVY, spaceAfter=10))
    story.append(Paragraph(f"Customer: {customer}", styles["CoverSub"]))
    story.append(Paragraph(f"Date: {TODAY_STR}  |  Maturity: {maturity}", styles["CoverSub"]))
    story.append(Paragraph(f"Confidentiality: {confidentiality}", styles["CoverSub"]))
    if prepared_by:
        story.append(Paragraph(f"Prepared by: {prepared_by}  |  Company: {company}", styles["CoverSub"]))
    story.append(Spacer(1, 1 * cm))

    summary_data = [
        ["Sell Price", "Base Cost", "Margin", "Mass (kg)"],
        [f"€ {sell:,.0f}", f"€ {base:,.0f}", f"€ {margin:,.0f} ({margin/sell*100:.1f}%)" if sell else "—",
         f"{total_mass:,.1f}"],
    ]
    tbl = Table(summary_data, colWidths=[4.2 * cm] * 4)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PDF_NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 12),
        ("BACKGROUND", (0, 1), (-1, 1), PDF_LIGHT),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFC9D0")),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(tbl)
    story.append(PageBreak())

    # ── Section 2: Executive Summary ─────────────────────────────────────────
    story.append(Paragraph("Executive Summary", styles["SectionHeading"]))

    mat = df["material_cost"].sum()
    moq_es = df["moq_excess_cost"].sum() if "moq_excess_cost" in df.columns else 0.0
    proc = df["process_cost"].sum()
    mach_es = df["machine_cost"].sum() if "machine_cost" in df.columns else 0.0
    lab_es = df["labour_cost"].sum() if "labour_cost" in df.columns else 0.0
    tool = df["tooling_cost"].sum() if "tooling_cost" in df.columns else 0.0
    energy_eur_es = df["energy_cost"].sum()
    kwh_total = df["energy_kwh"].sum() if "energy_kwh" in df.columns else energy_eur_es / 0.20
    rework = df["rework_cost"].sum() if "rework_cost" in df.columns else 0.0
    oh = df["overhead"].sum()
    marg = df["margin"].sum()

    pattern_es = df["pattern_cost"].sum() if "pattern_cost" in df.columns else 0.0

    kpi_headers = [["Item", "EUR", "% of Base", "Note"]]
    kpi_rows_pdf = [
        ("Material cost", mat, ""),
        ("MOQ excess cost", moq_es, "Min order qty premium"),
        ("Pattern / tooling NRE", pattern_es, "Casting pattern amortised per unit"),
        ("Process cost", proc, ""),
        ("  Machine cost", mach_es, "sub-total of Process"),
        ("  Labour cost", lab_es, "sub-total of Process"),
        (f"  Energy ({kwh_total:,.0f} kWh)", energy_eur_es, "@ 0.20 EUR/kWh"),
        ("  Tooling", tool, "sub-total of Process"),
        ("  Rework", rework, "sub-total of Process"),
        ("Overhead", oh, ""),
        ("Base cost", base, ""),
        ("Margin", marg, ""),
        ("Sell price", sell, ""),
    ]
    kpi_data = kpi_headers + [
        [n, f"€ {v:,.2f}", f"{v/base*100:.1f}%" if base else "—", nt]
        for n, v, nt in kpi_rows_pdf
    ]
    kpi_tbl = Table(kpi_data, colWidths=[6 * cm, 4 * cm, 3 * cm, 4 * cm])
    kpi_tbl.setStyle(_pdf_table_style(has_total_row=True, total_row_idx=len(kpi_data) - 3))
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.5 * cm))

    sub_df = _build_subsystem_table(df, bom)
    story.append(Paragraph("Subsystem Breakdown", styles["SectionHeading"]))
    sub_headers = [["Scope", "Cost EUR", "Share %", "Mass kg", "EUR/kg", "kWh"]]
    sub_rows = []
    if not sub_df.empty:
        for row in sub_df.itertuples():
            sub_rows.append([
                row.scope, f"€ {row.sell_eur:,.0f}",
                f"{row.share_pct:.1f}%", f"{row.mass_kg:,.1f}",
                f"€ {row.eur_per_kg:,.0f}", f"{row.kwh:,.0f}",
            ])
    sub_data = sub_headers + sub_rows
    sub_tbl = Table(sub_data, colWidths=[5 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    sub_tbl.setStyle(_pdf_table_style())
    story.append(sub_tbl)
    story.append(PageBreak())

    # ── Section 3: Cost Waterfall ─────────────────────────────────────────────
    story.append(Paragraph("Cost Waterfall", styles["SectionHeading"]))
    moq = df["moq_excess_cost"].sum() if "moq_excess_cost" in df.columns else 0.0
    pattern_wf2 = df["pattern_cost"].sum() if "pattern_cost" in df.columns else 0.0
    energy_eur_wf = df["energy_cost"].sum()
    kwh_wf = df["energy_kwh"].sum() if "energy_kwh" in df.columns else energy_eur_wf / 0.20
    wf_elements = [
        ("Material cost", df["material_cost"].sum()),
        ("MOQ excess cost", moq),
        ("Pattern / tooling NRE", pattern_wf2),
        ("Process cost", df["process_cost"].sum()),
        ("  Machine cost", df["machine_cost"].sum() if "machine_cost" in df.columns else 0.0),
        ("  Labour cost", df["labour_cost"].sum() if "labour_cost" in df.columns else 0.0),
        (f"  Energy cost ({kwh_wf:,.0f} kWh)", energy_eur_wf),
        ("  Tooling cost", df["tooling_cost"].sum() if "tooling_cost" in df.columns else 0.0),
        ("  Rework cost", df["rework_cost"].sum() if "rework_cost" in df.columns else 0.0),
        ("Overhead", df["overhead"].sum()),
        ("Base cost", df["base_cost"].sum()),
        ("Margin", df["margin"].sum()),
        ("Sell price", sell),
    ]
    wf_data = [["Cost Element", "EUR", "% of Sell Price"]] + [
        [name, f"€ {v:,.2f}", f"{v/sell*100:.1f}%" if sell else "—"]
        for name, v in wf_elements
    ]
    wf_tbl = Table(wf_data, colWidths=[6 * cm, 5 * cm, 5 * cm])
    wf_tbl.setStyle(_pdf_table_style(has_total_row=True, total_row_idx=len(wf_data) - 1))
    story.append(wf_tbl)
    story.append(PageBreak())

    # ── Section 4: Procurement Status ────────────────────────────────────────
    story.append(Paragraph("Procurement Status", styles["SectionHeading"]))

    today_ts = pd.Timestamp.today().normalize()
    if not quotes.empty and "valid_until" in quotes.columns:
        n_total = len(quotes)
        n_expired = (pd.to_datetime(quotes["valid_until"], errors="coerce") < today_ts).sum()
        story.append(Paragraph(
            f"Quote coverage: {n_total} quotes on file. {n_expired} expired.",
            styles["BodySmall"],
        ))

    proc_headers = [["Material", "Best Supplier", "Price EUR/kg", "Valid Until", "Status"]]
    proc_rows = []
    if not quotes.empty:
        for row in quotes.itertuples():
            try:
                vdt = pd.to_datetime(row.valid_until, errors="coerce")
                expired = bool(vdt < today_ts) if not pd.isnull(vdt) else False
                valid_str = vdt.strftime("%Y-%m-%d") if not pd.isnull(vdt) else ""
            except Exception:
                expired = False
                valid_str = str(row.valid_until)
            status_str = "(EXPIRED)" if expired else "ACTIVE"
            proc_rows.append([
                str(row.material_id), str(row.supplier),
                f"€ {row.price_eur_per_kg:,.3f}", valid_str, status_str,
            ])
    proc_data = proc_headers + proc_rows
    proc_tbl = Table(proc_data, colWidths=[4 * cm, 4.5 * cm, 3.5 * cm, 3 * cm, 2.5 * cm])
    proc_style = _pdf_table_style()
    # Mark expired rows
    for ri, row in enumerate(proc_rows, 1):
        if row[-1] == "(EXPIRED)":
            proc_style.add("BACKGROUND", (0, ri), (-1, ri), PDF_ORANGE)
    proc_tbl.setStyle(proc_style)
    story.append(proc_tbl)
    story.append(PageBreak())

    # ── Section 5: Energy & Carbon ───────────────────────────────────────────
    story.append(Paragraph("Energy & Carbon", styles["SectionHeading"]))
    story.append(Paragraph(
        "Scope 2 emissions based on purchased electricity for manufacturing operations, "
        f"using the EU average grid emission factor of {EMISSION_FACTOR} kg CO2e/kWh.",
        styles["BodySmall"],
    ))
    story.append(Spacer(1, 0.3 * cm))

    ec_headers = [["Scope", "kWh", "CO2e (kg)", "CO2e (t)", "Intensity kWh/EUR"]]
    ec_rows = []
    if not sub_df.empty:
        for row in sub_df.itertuples():
            co2_kg = row.kwh * EMISSION_FACTOR
            intensity = row.kwh / row.sell_eur if row.sell_eur > 0 else 0.0
            ec_rows.append([
                row.scope, f"{row.kwh:,.1f}", f"{co2_kg:,.1f}",
                f"{co2_kg/1000:,.3f}", f"{intensity:.4f}",
            ])
    ec_data = ec_headers + ec_rows
    ec_tbl = Table(ec_data, colWidths=[5.5 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm])
    ec_tbl.setStyle(_pdf_table_style())
    story.append(ec_tbl)
    story.append(Spacer(1, 0.3 * cm))

    total_co2 = kwh_total * EMISSION_FACTOR
    story.append(Paragraph(
        f"Total: <b>{kwh_total:,.0f} kWh</b>  |  <b>{total_co2:,.0f} kg CO2e</b>  |  {total_co2/1000:,.3f} t CO2e",
        styles["BodySmall"],
    ))
    story.append(Paragraph(
        "Note: Scope 2 boundary — purchased electricity for manufacturing. EU average grid 0.233 kg CO2e/kWh.",
        styles["NoteItalic"],
    ))
    story.append(PageBreak())

    # ── Section 6: Risk Register ─────────────────────────────────────────────
    story.append(Paragraph("Risk Register", styles["SectionHeading"]))
    risks = _build_risks(data)
    risk_headers = [["Category", "Title", "Severity", "Impact EUR", "Mitigation"]]
    risk_rows = [
        [r["Category"], r["Title"], r["Severity"],
         f"€ {r['Impact EUR']:,.0f}", r["Mitigation"]]
        for r in risks
    ]
    risk_data = risk_headers + risk_rows
    risk_tbl = Table(risk_data, colWidths=[3 * cm, 4.5 * cm, 2 * cm, 2.5 * cm, 5.5 * cm])
    risk_style = _pdf_table_style()
    sev_color_map = {"High": PDF_RED, "Medium": PDF_ORANGE, "Low": PDF_GREEN}
    for ri, row in enumerate(risk_rows, 1):
        sev = row[2]
        if sev in sev_color_map:
            risk_style.add("BACKGROUND", (2, ri), (2, ri), sev_color_map[sev])
            if sev == "High":
                risk_style.add("TEXTCOLOR", (2, ri), (2, ri), PDF_WHITE)
    risk_tbl.setStyle(risk_style)
    story.append(risk_tbl)

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT PAGE
# ══════════════════════════════════════════════════════════════════════════════

def main():
    inject_css()

    # ── Sidebar ───────────────────────────────────────────────────────────────
    st.sidebar.header("Report Settings")
    company_name = st.sidebar.text_input("Company name", value="Cost Forge Engineering")
    prepared_by = st.sidebar.text_input("Prepared by", value="")
    confidentiality = st.sidebar.selectbox(
        "Confidentiality",
        ["Internal", "Commercial in Confidence", "Restricted"],
        index=1,
    )

    # ── Header ────────────────────────────────────────────────────────────────
    home_button()
    data = _load_all()
    meta = data["meta"]
    df = data["df"]
    bom = data["bom"]
    quotes = data["quotes"]

    proj_name = meta.get("name", "Unnamed Project")
    customer = meta.get("customer", "")
    estimate_ref = meta.get("estimate_ref", "")
    maturity = meta.get("maturity", "Budget (±15%)")

    page_header(
        title="Stakeholder Package",
        icon="📦",
        caption="Full project workflow — Excel workbook + PDF report for stakeholder communication",
        project=proj_name,
        maturity=maturity,
    )

    # ── Project meta fields in sidebar ───────────────────────────────────────
    st.sidebar.divider()
    st.sidebar.header("Project Details")
    st.sidebar.caption("These fields populate the Cover sheet of the report.")
    customer_in = st.sidebar.text_input("Customer", value=customer)
    estimate_ref_in = st.sidebar.text_input("Estimate reference", value=estimate_ref,
                                             placeholder="e.g. CE-2024-047 Rev B")
    contract_value_in = st.sidebar.number_input(
        "Contract value (EUR, if awarded)", min_value=0.0, value=float(meta.get("contract_value", 0.0)),
        step=1000.0, format="%.0f"
    )
    if st.sidebar.button("💾 Save project details", use_container_width=True):
        from utils.project import save_project_meta
        save_project_meta(
            customer=customer_in,
            estimate_ref=estimate_ref_in,
            contract_value=contract_value_in,
        )
        st.sidebar.success("Saved.")
        st.cache_data.clear()
        st.rerun()

    # Merge saved fields into data dict meta for this run
    data["meta"] = {**meta, "customer": customer_in, "estimate_ref": estimate_ref_in,
                    "contract_value": contract_value_in}

    # ── KPI row ───────────────────────────────────────────────────────────────
    sell = df["total_cost"].sum()
    base = df["base_cost"].sum()
    margin = df["margin"].sum()
    total_mass = df["mass_kg"].sum() if "mass_kg" in df.columns else 0.0
    kwh_total = df["energy_kwh"].sum() if "energy_kwh" in df.columns else df["energy_cost"].sum() / 0.20
    moq_total = df["moq_excess_cost"].sum() if "moq_excess_cost" in df.columns else 0.0
    proc_total = df["process_cost"].sum()
    co2e_kg = kwh_total * EMISSION_FACTOR

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Sell Price", fmt(sell, 0))
    k2.metric("Base Cost", fmt(base, 0))
    k3.metric("Margin", f"{margin/sell*100:.1f}%" if sell else "—",
              delta=f"{fmt(margin, 0)}", delta_color="normal")
    k4.metric("Total Mass", f"{total_mass:,.0f} kg")
    k5.metric("Energy", f"{kwh_total:,.0f} kWh",
              delta=f"{co2e_kg/1000:,.2f} t CO₂e", delta_color="off")
    k6.metric("MOQ Excess", fmt(moq_total, 0),
              delta="⚠ larger than material" if moq_total > df["material_cost"].sum() else None,
              delta_color="inverse")

    st.divider()

    # ── File naming ───────────────────────────────────────────────────────────
    proj_slug = re.sub(r"[^\w]", "_", proj_name).lower()[:40]
    excel_name = f"stakeholder_package_{proj_slug}_{TODAY_STR}.xlsx"
    pdf_name = f"stakeholder_report_{proj_slug}_{TODAY_STR}.pdf"

    # ── Download buttons ──────────────────────────────────────────────────────
    col1, col2 = st.columns(2)

    with col1:
        with st.spinner("Building Excel workbook…"):
            xl_bytes = build_excel(data, company_name, prepared_by, confidentiality)
        st.download_button(
            label="📥 Download Stakeholder Package (Excel)",
            data=xl_bytes,
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

    with col2:
        with st.spinner("Building PDF report…"):
            pdf_bytes = build_pdf(data, company_name, prepared_by, confidentiality)
        st.download_button(
            label="📄 Download Stakeholder Report (PDF)",
            data=pdf_bytes,
            file_name=pdf_name,
            mime="application/pdf",
            use_container_width=True,
        )

    st.divider()

    # ── Rich on-screen preview ────────────────────────────────────────────────
    st.subheader("📊 Preview — What's in the package")

    tab_cost, tab_subsys, tab_procure, tab_energy, tab_risks, tab_contents = st.tabs([
        "💶 Cost Breakdown",
        "🔩 Subsystem Split",
        "🛒 Procurement",
        "⚡ Energy & Carbon",
        "⚠️ Risks",
        "📋 Contents",
    ])

    # ── Tab 1: Cost Breakdown ─────────────────────────────────────────────────
    with tab_cost:
        mat = df["material_cost"].sum()
        mach = df["machine_cost"].sum() if "machine_cost" in df.columns else 0.0
        lab = df["labour_cost"].sum() if "labour_cost" in df.columns else 0.0
        tool = df["tooling_cost"].sum() if "tooling_cost" in df.columns else 0.0
        energy_eur = df["energy_cost"].sum()
        rework = df["rework_cost"].sum() if "rework_cost" in df.columns else 0.0
        oh = df["overhead"].sum()
        marg_v = df["margin"].sum()

        pattern_tab = df["pattern_cost"].sum() if "pattern_cost" in df.columns else 0.0
        cost_rows = [
            {"Cost Element": "Material cost",          "EUR": mat,         "% of Base": f"{mat/base*100:.1f}%",         "Note": ""},
            {"Cost Element": "MOQ excess cost",        "EUR": moq_total,   "% of Base": f"{moq_total/base*100:.1f}%",   "Note": "⚠ min order qty premium — often > material cost"},
            {"Cost Element": "Pattern / tooling NRE",  "EUR": pattern_tab, "% of Base": f"{pattern_tab/base*100:.1f}%", "Note": "Casting pattern amortised per unit"},
            {"Cost Element": "Process cost",           "EUR": proc_total,  "% of Base": f"{proc_total/base*100:.1f}%",  "Note": "dominant cost driver"},
            {"Cost Element": "  Machine cost",    "EUR": mach,       "% of Base": f"{mach/base*100:.1f}%",  "Note": "sub-total of Process"},
            {"Cost Element": "  Labour cost",     "EUR": lab,        "% of Base": f"{lab/base*100:.1f}%",   "Note": "sub-total of Process"},
            {"Cost Element": f"  Energy ({kwh_total:,.0f} kWh)", "EUR": energy_eur, "% of Base": f"{energy_eur/base*100:.1f}%", "Note": "@ 0.20 EUR/kWh"},
            {"Cost Element": "  Tooling cost",    "EUR": tool,       "% of Base": f"{tool/base*100:.1f}%",  "Note": "sub-total of Process"},
            {"Cost Element": "  Rework cost",     "EUR": rework,     "% of Base": f"{rework/base*100:.1f}%","Note": "sub-total of Process"},
            {"Cost Element": "Overhead",          "EUR": oh,         "% of Base": f"{oh/base*100:.1f}%",    "Note": ""},
            {"Cost Element": "BASE COST",         "EUR": base,       "% of Base": "100.0%",                 "Note": ""},
            {"Cost Element": "Margin",            "EUR": marg_v,     "% of Base": f"{marg_v/base*100:.1f}%","Note": f"margin on sell = {marg_v/sell*100:.1f}%"},
            {"Cost Element": "SELL PRICE",        "EUR": sell,       "% of Base": f"{sell/base*100:.1f}%",  "Note": ""},
        ]
        cost_df = pd.DataFrame(cost_rows)
        cost_df["EUR"] = cost_df["EUR"].map(lambda v: f"€ {v:,.2f}")
        st.dataframe(cost_df, use_container_width=True, hide_index=True)

        if moq_total > mat:
            st.warning(
                f"⚠️ **MOQ excess cost (€ {moq_total:,.0f}) is larger than material cost (€ {mat:,.0f})**. "
                "This is a significant cost driver — review BOM quantities and negotiate blanket orders with suppliers."
            )

    # ── Tab 2: Subsystem Split ────────────────────────────────────────────────
    with tab_subsys:
        sub_df = _build_subsystem_table(df, bom)
        if sub_df.empty:
            st.info("No BOM lines loaded.")
        else:
            disp = sub_df[["scope", "lines", "mass_kg", "material_eur", "process_eur",
                           "sell_eur", "share_pct", "eur_per_kg", "kwh", "co2e_kg"]].copy()
            disp.columns = ["Scope", "Lines", "Mass (kg)", "Material (EUR)", "Process (EUR)",
                            "Sell (EUR)", "Share %", "EUR/kg", "kWh", "CO₂e kg"]
            for c in ["Mass (kg)", "Material (EUR)", "Process (EUR)", "Sell (EUR)", "EUR/kg", "kWh", "CO₂e kg"]:
                disp[c] = pd.to_numeric(disp[c]).map(lambda v: f"{v:,.1f}")
            disp["Share %"] = pd.to_numeric(sub_df["share_pct"]).map(lambda v: f"{v:.1f}%")
            st.dataframe(disp, use_container_width=True, hide_index=True)
            st.caption(f"Total: {len(bom)} BOM lines  |  {bom['material_id'].nunique() if 'material_id' in bom.columns else '?'} unique materials  |  {total_mass:,.1f} kg")

    # ── Tab 3: Procurement ────────────────────────────────────────────────────
    with tab_procure:
        if quotes.empty:
            st.warning("No supplier quotes on file. Add quotes on the **Procurement** page to populate this section.")
        else:
            today_ts = pd.Timestamp.today().normalize()
            q_disp = quotes.copy()
            if "valid_until" in q_disp.columns:
                q_disp["Status"] = q_disp["valid_until"].apply(
                    lambda d: "EXPIRED" if pd.to_datetime(d, errors="coerce") < today_ts else "ACTIVE"
                )
            n_exp = (q_disp.get("Status", pd.Series(dtype=str)) == "EXPIRED").sum()
            if n_exp:
                st.error(f"⚠ {n_exp} quote(s) expired — refresh before sending to customer.")
            else:
                st.success(f"✅ All {len(quotes)} quotes are within validity.")
            st.dataframe(q_disp, use_container_width=True, hide_index=True)

    # ── Tab 4: Energy & Carbon ────────────────────────────────────────────────
    with tab_energy:
        co2e_t = co2e_kg / 1000
        c1, c2, c3 = st.columns(3)
        c1.metric("Total Energy", f"{kwh_total:,.0f} kWh")
        c2.metric("CO₂e (EU grid)", f"{co2e_kg:,.0f} kg", delta=f"{co2e_t:.2f} t", delta_color="off")
        c3.metric("Energy Intensity", f"{kwh_total/sell:.3f} kWh/EUR" if sell else "—")

        st.caption("Scope 2 emissions — purchased electricity for manufacturing. EU average grid 0.233 kg CO2e/kWh.")

        sub_df2 = _build_subsystem_table(df, bom)
        if not sub_df2.empty:
            ec_disp = sub_df2[["scope", "kwh", "co2e_kg"]].copy()
            ec_disp["co2e_t"] = ec_disp["co2e_kg"] / 1000
            ec_disp["kWh/EUR"] = sub_df2.apply(
                lambda r: r["kwh"] / r["sell_eur"] if r["sell_eur"] > 0 else 0.0, axis=1
            )
            ec_disp.columns = ["Scope", "kWh", "CO₂e (kg)", "CO₂e (t)", "kWh/EUR intensity"]
            for c in ["kWh", "CO₂e (kg)", "CO₂e (t)"]:
                ec_disp[c] = pd.to_numeric(ec_disp[c]).map(lambda v: f"{v:,.1f}")
            ec_disp["kWh/EUR intensity"] = ec_disp["kWh/EUR intensity"].map(lambda v: f"{v:.4f}")
            st.dataframe(ec_disp, use_container_width=True, hide_index=True)

    # ── Tab 5: Risks ──────────────────────────────────────────────────────────
    with tab_risks:
        risks = _build_risks(data)
        if not risks:
            st.success("No risks identified.")
        else:
            risk_disp = pd.DataFrame(risks)
            if "Impact EUR" in risk_disp.columns:
                risk_disp["Impact EUR"] = risk_disp["Impact EUR"].map(lambda v: f"€ {v:,.0f}")
            st.dataframe(risk_disp, use_container_width=True, hide_index=True)

            high_risks = [r for r in risks if r.get("Severity") == "High"]
            if high_risks:
                st.error(f"🔴 {len(high_risks)} HIGH severity risk(s) require attention before submission.")

    # ── Tab 6: Contents ───────────────────────────────────────────────────────
    with tab_contents:
        # Data completeness checklist
        nre = data.get("nre", pd.DataFrame())
        milestones = data.get("milestones", pd.DataFrame())
        transport = data.get("transport", pd.DataFrame())
        india_lc = data.get("india_lc", pd.DataFrame())
        change_orders = data.get("change_orders", pd.DataFrame())

        checks = [
            ("Project name set", bool(proj_name and proj_name != "Unnamed Project")),
            ("Customer set", bool(data["meta"].get("customer", ""))),
            ("Estimate reference set", bool(data["meta"].get("estimate_ref", ""))),
            (f"BOM loaded ({len(bom)} lines)", not bom.empty),
            (f"Quotes on file ({len(quotes)})", not quotes.empty),
            (f"India LC data ({len(india_lc)} rows)", not india_lc.empty),
            (f"NRE / non-recurring costs ({len(nre)} rows)", not (isinstance(nre, pd.DataFrame) and nre.empty)),
            (f"Milestones ({len(milestones)} rows)", not (isinstance(milestones, pd.DataFrame) and milestones.empty)),
            (f"Transport quotes ({len(transport)} rows)", not (isinstance(transport, pd.DataFrame) and transport.empty)),
            (f"Change orders ({len(change_orders)} rows)", not (isinstance(change_orders, pd.DataFrame) and change_orders.empty)),
        ]

        st.markdown("#### Package data completeness")
        for label, ok in checks:
            icon = "✅" if ok else "⬜"
            st.markdown(f"{icon} {label}")

        empty_sections = [label for label, ok in checks if not ok]
        if empty_sections:
            st.info(
                f"**{len(empty_sections)} section(s) not yet populated.** "
                "Empty sections will show placeholder messages in the Excel/PDF. "
                "Use the respective pages (Procurement, India Local Content, NRE, etc.) to fill in data."
            )

        st.divider()
        st.markdown("#### Package contents")
        preview_data = {
            "Excel Sheet": [
                "1. COVER",
                "2. EXECUTIVE SUMMARY",
                "3. COST WATERFALL",
                "4. SUBSYSTEM BREAKDOWN",
                "5. BOM DETAIL",
                "6. PROCUREMENT",
                "7. ENERGY & CARBON",
                "8. INDIA LOCAL CONTENT",
                "9. RISK & ACTIONS",
            ],
            "Description": [
                "Project metadata, KPIs, scope statement",
                "Cost KPIs with process sub-breakdown + subsystem table",
                "Waterfall with MOQ, machine, labour, energy, tooling, rework",
                "14 subsystem scopes: cost, mass, share, kWh, CO2e",
                "All BOM lines with full cost breakdown",
                "Supplier quotes, lead times, expiry status",
                "Energy by subsystem + carbon footprint (0.233 kg CO2e/kWh)",
                "India local content declarations and IC fraction",
                "Auto-generated risks + risk register (severity, impact, mitigation)",
            ],
            "PDF Section": [
                "§1 Cover Page",
                "§2 Executive Summary",
                "§3 Cost Waterfall",
                "§2 Subsystem Table",
                "(Excel only)",
                "§4 Procurement Status",
                "§5 Energy & Carbon",
                "(Excel only)",
                "§6 Risk Register",
            ],
        }
        st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)


guard(main)
